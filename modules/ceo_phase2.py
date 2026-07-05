from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from modules.business_intelligence import (
    AFFILIATE_MATCH_FIELDS,
    AI_RECOMMENDATION_FIELDS,
    EXECUTION_FIELDS,
    MONEY_FIELDS,
    TREND_MOMENTUM_FIELDS,
    build_affiliate_match,
    build_ai_recommendations,
    build_execution_tracker,
    build_money_score,
    build_opportunity_breakdown,
    build_seo_difficulty,
    build_trend_momentum,
    infer_intent,
)
from modules.performance_tracking import DATA_DIR, numeric, print_workbook_report, read_csv, safe_save_workbook, update_master_workbook, write_csv, write_json


CEO_FIELDS = ["metric", "value", "notes"]
ROI_FIELDS = [
    "rank",
    "topic",
    "slug",
    "expected_revenue",
    "expected_clicks",
    "expected_conversion",
    "expected_affiliate_revenue",
    "estimated_writing_time",
    "estimated_video_production_time",
    "estimated_social_time",
    "estimated_total_cost",
    "estimated_profit",
    "roi",
    "roi_per_hour",
    "priority_roi_score",
]
DECISION_FIELDS = ["topic", "slug", "action", "reason", "confidence", "expected_revenue", "expected_traffic", "deadline", "risk", "priority"]
PREDICTION_FIELDS = [
    "topic",
    "slug",
    "traffic_7d",
    "traffic_30d",
    "revenue_30d",
    "revenue_90d",
    "trend_growth",
    "trend_decline",
    "topic_lifetime",
    "competition_increase",
    "seo_difficulty_change",
    "opportunity_window",
]
ADVANCED_MOMENTUM_FIELDS = [
    "topic",
    "slug",
    "velocity",
    "acceleration",
    "growth_rate",
    "decay_rate",
    "stability_score",
    "volatility_score",
    "days_since_first_seen",
    "days_since_last_growth",
    "peak_rank",
    "average_rank",
    "trend_lifetime",
    "trend_confidence",
    "momentum_grade",
    "momentum_recommendation",
]
CALENDAR_FIELDS = [
    "date",
    "time",
    "topic",
    "article",
    "video",
    "short",
    "facebook",
    "linkedin",
    "reddit",
    "quora",
    "x",
    "newsletter",
    "priority",
    "expected_revenue",
    "status",
]
COMPETITOR_PHASE2_FIELDS = [
    "topic",
    "competitor_url",
    "word_count",
    "images",
    "videos",
    "headings",
    "faqs",
    "schema",
    "internal_links",
    "external_links",
    "affiliate_programs",
    "cta_count",
    "update_frequency",
    "content_freshness",
    "estimated_traffic",
    "estimated_authority",
    "missing_topics",
    "gap_score",
    "opportunity_score",
]
CONTENT_GAP_FIELDS = ["topic", "slug", "gap_type", "recommendation", "priority", "estimated_revenue", "difficulty", "source"]
AFFILIATE_INTEL_FIELDS = [
    "topic",
    "slug",
    "Amazon",
    "PartnerStack",
    "Impact",
    "CJ",
    "Rakuten",
    "ShareASale",
    "Awin",
    "Digistore24",
    "ClickBank",
    "Direct SaaS",
    "recurring_percent",
    "cookie_length",
    "approval_difficulty",
    "commission",
    "epc",
    "expected_revenue",
    "best_match_score",
    "affiliate_priority",
]
ARTICLE_PIPELINE_FIELDS = [
    "topic",
    "slug",
    "research",
    "outline",
    "writing",
    "seo",
    "images",
    "video",
    "thumbnail",
    "social",
    "publish",
    "index",
    "backlinks",
    "refresh_schedule",
    "status",
    "completion_percent",
    "owner",
]
EXECUTION_PHASE2_FIELDS = EXECUTION_FIELDS + ["completion_percent", "time_spent", "money_earned", "revenue_forecast", "next_action", "waiting_reason", "blocked_reason", "missed_deadline"]
AI_SUMMARY_FIELDS = ["section", "summary"]


def today_iso() -> str:
    return date.today().isoformat()


def money_by_slug() -> dict[str, dict[str, Any]]:
    return {row.get("slug"): row for row in build_money_score()}


def seo_by_slug() -> dict[str, dict[str, Any]]:
    return {row.get("slug"): row for row in build_seo_difficulty()}


def build_roi_analysis() -> list[dict[str, Any]]:
    money = build_money_score()
    seo = seo_by_slug()
    rows = []
    for row in money:
        slug = row.get("slug", "")
        intent = infer_intent(str(row.get("topic")))
        expected_revenue = numeric(row.get("estimated_monthly_revenue"))
        expected_clicks = numeric(row.get("estimated_affiliate_clicks"))
        expected_conversion = numeric(row.get("estimated_conversion_rate"))
        writing_time = 2.5 if intent in {"review", "comparison", "best list"} else 1.5
        video_time = 1.5 if expected_revenue >= 10 or intent in {"review", "comparison"} else 0.5
        social_time = 0.4 if expected_revenue >= 5 else 0.2
        total_hours = writing_time + video_time + social_time
        total_cost = round(total_hours * 18, 2)
        profit = round(expected_revenue - total_cost, 2)
        roi = round(profit / max(total_cost, 1), 2)
        roi_per_hour = round(profit / max(total_hours, 0.1), 2)
        difficulty_penalty = numeric(seo.get(slug, {}).get("keyword_difficulty_estimate")) * 0.2
        priority = round(numeric(row.get("money_score")) + roi_per_hour * 0.4 - difficulty_penalty, 1)
        rows.append(
            {
                "rank": 0,
                "topic": row.get("topic", ""),
                "slug": slug,
                "expected_revenue": expected_revenue,
                "expected_clicks": expected_clicks,
                "expected_conversion": expected_conversion,
                "expected_affiliate_revenue": expected_revenue,
                "estimated_writing_time": writing_time,
                "estimated_video_production_time": video_time,
                "estimated_social_time": social_time,
                "estimated_total_cost": total_cost,
                "estimated_profit": profit,
                "roi": roi,
                "roi_per_hour": roi_per_hour,
                "priority_roi_score": priority,
            }
        )
    rows.sort(key=lambda item: numeric(item.get("roi_per_hour")), reverse=True)
    for index, row in enumerate(rows, 1):
        row["rank"] = index
    return rows


def build_ai_decisions() -> list[dict[str, Any]]:
    recommendations = build_ai_recommendations()
    money = money_by_slug()
    seo = seo_by_slug()
    rows = []
    for row in recommendations:
        slug = row.get("slug", "")
        m = money.get(slug, {})
        s = seo.get(slug, {})
        base_action = str(row.get("final_action", "Watch"))
        action = map_action(base_action, row, m, s)
        priority = "P1" if action in {"WRITE NOW", "INDEX NOW"} else "P2" if action in {"CREATE VIDEO", "CREATE COMPARISON PAGE", "REFRESH ARTICLE"} else "P3" if action != "IGNORE" else "P5"
        rows.append(
            {
                "topic": row.get("topic", ""),
                "slug": slug,
                "action": action,
                "reason": row.get("reason", ""),
                "confidence": m.get("revenue_confidence", ""),
                "expected_revenue": m.get("estimated_monthly_revenue", ""),
                "expected_traffic": m.get("estimated_monthly_traffic", ""),
                "deadline": deadline_for_priority(priority),
                "risk": row.get("risk") or s.get("serp_risk", ""),
                "priority": priority,
            }
        )
    return rows


def map_action(base_action: str, row: dict[str, Any], money: dict[str, Any], seo: dict[str, Any]) -> str:
    intent = infer_intent(str(row.get("topic")))
    revenue = numeric(money.get("estimated_monthly_revenue"))
    if base_action == "Write Now":
        return "CREATE COMPARISON PAGE" if intent == "comparison" else "WRITE NOW"
    if base_action == "Write Today":
        return "CREATE LANDING PAGE" if intent == "pricing" and revenue >= 15 else "WRITE LATER"
    if base_action == "Refresh Existing":
        return "REFRESH ARTICLE"
    if base_action == "Make Video":
        return "CREATE VIDEO"
    if numeric(seo.get("ranking_opportunity")) >= 70 and revenue >= 5:
        return "INDEX NOW"
    if revenue < 1 and row.get("risk") == "High":
        return "IGNORE"
    return "WAIT"


def deadline_for_priority(priority: str) -> str:
    days = {"P1": 0, "P2": 2, "P3": 7, "P5": 30}.get(priority, 7)
    return (date.today() + timedelta(days=days)).isoformat()


def build_predictions() -> list[dict[str, Any]]:
    money = build_money_score()
    momentum = {row.get("slug"): row for row in build_advanced_momentum()}
    seo = seo_by_slug()
    rows = []
    for row in money:
        slug = row.get("slug", "")
        m = momentum.get(slug, {})
        s = seo.get(slug, {})
        traffic = numeric(row.get("estimated_monthly_traffic"))
        revenue = numeric(row.get("estimated_monthly_revenue"))
        growth = max(-0.3, min(1.5, numeric(m.get("growth_rate")) / 100))
        decline = numeric(m.get("decay_rate")) / 100
        rows.append(
            {
                "topic": row.get("topic", ""),
                "slug": slug,
                "traffic_7d": int(traffic / 4 * (1 + growth)),
                "traffic_30d": int(traffic * (1 + growth - decline * 0.3)),
                "revenue_30d": round(revenue * (1 + growth - decline * 0.3), 2),
                "revenue_90d": round(revenue * 3 * (1 + growth * 0.5), 2),
                "trend_growth": round(growth * 100, 1),
                "trend_decline": round(decline * 100, 1),
                "topic_lifetime": m.get("trend_lifetime", ""),
                "competition_increase": round(numeric(s.get("keyword_difficulty_estimate")) * 0.08, 1),
                "seo_difficulty_change": "Increasing" if numeric(s.get("keyword_difficulty_estimate")) > 65 else "Stable",
                "opportunity_window": opportunity_window(m, s),
            }
        )
    return rows


def build_advanced_momentum() -> list[dict[str, Any]]:
    base = build_trend_momentum()
    history = read_csv(DATA_DIR / "hottrend_topic_history.csv")
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in history:
        grouped[row.get("slug") or row.get("topic")].append(row)
    rows = []
    for item in base:
        slug = item.get("slug", "")
        rows_for_slug = grouped.get(slug, [])
        scores = [numeric(row.get("total_score") or row.get("score_today")) for row in rows_for_slug] or [numeric(item.get("score_today"))]
        ranks = [numeric(row.get("rank_today") or row.get("rank")) for row in rows_for_slug if numeric(row.get("rank_today") or row.get("rank"))] or [numeric(item.get("rank_today"))]
        velocity = round((scores[-1] - scores[0]) / max(len(scores) - 1, 1), 2)
        previous_velocity = round((scores[-2] - scores[0]) / max(len(scores) - 2, 1), 2) if len(scores) > 2 else 0
        acceleration = round(velocity - previous_velocity, 2)
        volatility = round(max(scores) - min(scores), 1)
        stability = round(max(0, 100 - volatility), 1)
        growth_rate = round((scores[-1] - scores[0]) / max(scores[0], 1) * 100, 1)
        decay_rate = round(max(0, (max(scores) - scores[-1]) / max(max(scores), 1) * 100), 1)
        rows.append(
            {
                "topic": item.get("topic", ""),
                "slug": slug,
                "velocity": velocity,
                "acceleration": acceleration,
                "growth_rate": growth_rate,
                "decay_rate": decay_rate,
                "stability_score": stability,
                "volatility_score": volatility,
                "days_since_first_seen": item.get("trend_age_days", ""),
                "days_since_last_growth": 0 if growth_rate > 0 else item.get("trend_age_days", ""),
                "peak_rank": int(min(ranks)),
                "average_rank": round(sum(ranks) / max(len(ranks), 1), 1),
                "trend_lifetime": item.get("trend_age_days", ""),
                "trend_confidence": round((stability + numeric(item.get("score_today"))) / 2, 1),
                "momentum_grade": grade_momentum(growth_rate, stability),
                "momentum_recommendation": recommend_momentum(growth_rate, stability),
            }
        )
    return rows


def grade_momentum(growth_rate: float, stability: float) -> str:
    if growth_rate >= 15 and stability >= 70:
        return "A"
    if growth_rate >= 5:
        return "B"
    if growth_rate >= -5:
        return "C"
    return "D"


def recommend_momentum(growth_rate: float, stability: float) -> str:
    if growth_rate >= 15 and stability >= 70:
        return "Act this week"
    if growth_rate >= 5:
        return "Monitor and prepare"
    if growth_rate < -10:
        return "Avoid unless strong buyer intent"
    return "Watch"


def opportunity_window(momentum: dict[str, Any], seo: dict[str, Any]) -> str:
    if str(momentum.get("momentum_grade")) in {"A", "B"} and numeric(seo.get("keyword_difficulty_estimate")) < 70:
        return "Now to 14 days"
    if numeric(seo.get("keyword_difficulty_estimate")) >= 70:
        return "Long-tail support only"
    return "Next 30 days"


def build_content_calendar() -> list[dict[str, Any]]:
    decisions = build_ai_decisions()
    rows = []
    start = datetime.combine(date.today(), datetime.min.time()).replace(hour=9)
    day_offset = 0
    for index, decision in enumerate(decisions[:30]):
        scheduled = start + timedelta(days=day_offset, hours=(index % 3) * 3)
        if index and index % 3 == 0:
            day_offset += 1
        action = decision.get("action", "")
        rows.append(
            {
                "date": scheduled.date().isoformat(),
                "time": scheduled.strftime("%H:%M"),
                "topic": decision.get("topic", ""),
                "article": "Yes" if action in {"WRITE NOW", "WRITE LATER", "CREATE COMPARISON PAGE", "CREATE LANDING PAGE"} else "No",
                "video": "Yes" if action == "CREATE VIDEO" else "No",
                "short": "Yes" if action in {"CREATE VIDEO", "CREATE SHORT"} else "No",
                "facebook": "Draft" if decision.get("priority") in {"P1", "P2"} else "No",
                "linkedin": "Draft" if decision.get("priority") in {"P1", "P2"} else "No",
                "reddit": "Draft" if decision.get("priority") == "P1" else "No",
                "quora": "Draft" if decision.get("priority") == "P1" else "No",
                "x": "Draft" if decision.get("priority") in {"P1", "P2"} else "No",
                "newsletter": "Draft" if decision.get("priority") == "P1" else "No",
                "priority": decision.get("priority", ""),
                "expected_revenue": decision.get("expected_revenue", ""),
                "status": "Planned",
            }
        )
    return rows


def build_competitor_phase2() -> list[dict[str, Any]]:
    rows = read_csv(DATA_DIR / "competitor_intelligence.csv") or read_csv(DATA_DIR / "competitor_watch.csv")
    if not rows:
        rows = [{"topic": row.get("topic", ""), "competitor_url": "Needs manual input"} for row in build_opportunity_breakdown()[:20]]
    output = []
    for row in rows:
        topic = row.get("topic") or row.get("target_keyword") or "Needs manual input"
        strength = numeric(row.get("estimated_strength") or row.get("difficulty"), 50)
        freshness = 55 if row.get("content_freshness") else 35
        output.append(
            {
                "topic": topic,
                "competitor_url": row.get("competitor_url", "Needs manual input"),
                "word_count": row.get("word_count", "Needs manual input"),
                "images": row.get("images", "Needs manual input"),
                "videos": row.get("videos", "Needs manual input"),
                "headings": row.get("headings", "Needs manual input"),
                "faqs": row.get("faqs", "Needs manual input"),
                "schema": row.get("schema", "Needs manual input"),
                "internal_links": row.get("internal_links", "Needs manual input"),
                "external_links": row.get("external_links", "Needs manual input"),
                "affiliate_programs": row.get("affiliate_programs", "Needs manual input"),
                "cta_count": row.get("cta_count", "Needs manual input"),
                "update_frequency": row.get("update_frequency", "Needs manual input"),
                "content_freshness": row.get("content_freshness", "Needs manual input"),
                "estimated_traffic": row.get("estimated_traffic", ""),
                "estimated_authority": strength,
                "missing_topics": row.get("missing_topics") or row.get("content_gap") or "Needs manual input",
                "gap_score": round(max(0, 100 - strength) * 0.6 + freshness * 0.4, 1),
                "opportunity_score": round(max(0, 100 - strength) * 0.5 + freshness * 0.5, 1),
            }
        )
    return output


def build_content_gap() -> list[dict[str, Any]]:
    opportunity = build_opportunity_breakdown()
    existing_slugs = {row.get("slug") for row in read_csv(DATA_DIR / "content_lifecycle.csv")}
    rows = []
    for row in opportunity:
        slug = row.get("slug", "")
        intent = row.get("intent", "")
        if slug not in existing_slugs:
            gap_type = f"Missing {intent} page"
        elif intent in {"review", "comparison", "pricing", "alternative"} and numeric(row.get("total_score")) >= 70:
            gap_type = "Topic needing update"
        else:
            continue
        rows.append(
            {
                "topic": row.get("topic", ""),
                "slug": slug,
                "gap_type": gap_type,
                "recommendation": f"Create or refresh a {intent} asset with FAQ, comparison table, pricing, and video placeholder.",
                "priority": "High" if numeric(row.get("total_score")) >= 75 else "Medium",
                "estimated_revenue": money_by_slug().get(slug, {}).get("estimated_monthly_revenue", ""),
                "difficulty": seo_by_slug().get(slug, {}).get("keyword_difficulty_estimate", ""),
                "source": "topic_scores",
            }
        )
    return rows


def build_affiliate_intelligence() -> list[dict[str, Any]]:
    affiliate = build_affiliate_match()
    money = money_by_slug()
    rows = []
    for row in affiliate:
        slug = row.get("slug", "")
        topic = str(row.get("topic"))
        direct = "Likely" if any(term in topic.lower() for term in ("saas", "software", "tool", "seo", "automation")) else "Check"
        score = numeric(row.get("program_confidence")) + numeric(money.get(slug, {}).get("estimated_monthly_revenue")) * 0.2
        rows.append(
            {
                "topic": topic,
                "slug": slug,
                "Amazon": "Check" if "hardware" in topic.lower() else "No",
                "PartnerStack": row.get("PartnerStack", "Check"),
                "Impact": row.get("Impact", "Check"),
                "CJ": row.get("CJ", "Check"),
                "Rakuten": "Check",
                "ShareASale": row.get("ShareASale", "Check"),
                "Awin": "Check",
                "Digistore24": "Check",
                "ClickBank": "Check",
                "Direct SaaS": direct,
                "recurring_percent": 20 if direct == "Likely" else 0,
                "cookie_length": "30-90 days, verify terms",
                "approval_difficulty": "Medium" if direct == "Likely" else "Unknown",
                "commission": row.get("commission_estimate", ""),
                "epc": round(numeric(row.get("commission_estimate")) * 0.03, 2),
                "expected_revenue": money.get(slug, {}).get("estimated_monthly_revenue", ""),
                "best_match_score": round(score, 1),
                "affiliate_priority": row.get("affiliate_priority", ""),
            }
        )
    return rows


def build_article_pipeline() -> list[dict[str, Any]]:
    decisions = build_ai_decisions()
    execution = {row.get("slug"): row for row in build_execution_tracker()}
    rows = []
    for row in decisions:
        slug = row.get("slug", "")
        ex = execution.get(slug, {})
        action = row.get("action", "")
        completion = pipeline_completion(ex)
        rows.append(
            {
                "topic": row.get("topic", ""),
                "slug": slug,
                "research": "Pending",
                "outline": "Pending",
                "writing": "Pending",
                "seo": "Pending",
                "images": "Pending",
                "video": "Pending" if action in {"CREATE VIDEO", "WRITE NOW"} else "Optional",
                "thumbnail": "Pending" if action == "CREATE VIDEO" else "Optional",
                "social": "Pending" if row.get("priority") in {"P1", "P2"} else "Optional",
                "publish": "Pending",
                "index": "Pending",
                "backlinks": "Pending",
                "refresh_schedule": (date.today() + timedelta(days=90)).isoformat(),
                "status": "Pending",
                "completion_percent": completion,
                "owner": "User / Bot",
            }
        )
    return rows


def pipeline_completion(execution: dict[str, Any]) -> int:
    done = sum(1 for field in ("article_created", "website_published", "youtube_uploaded", "indexed_google", "revenue_checked") if execution.get(field) == "Yes")
    return int(done / 5 * 100)


def build_execution_phase2() -> list[dict[str, Any]]:
    execution = build_execution_tracker()
    money = money_by_slug()
    decisions = {row.get("slug"): row for row in build_ai_decisions()}
    rows = []
    for row in execution:
        slug = row.get("slug", "")
        decision = decisions.get(slug, {})
        missed = "Yes" if decision.get("deadline") and str(decision.get("deadline")) < today_iso() and row.get("website_published") != "Yes" else "No"
        enriched = dict(row)
        enriched.update(
            {
                "completion_percent": pipeline_completion(row),
                "time_spent": "",
                "money_earned": "",
                "revenue_forecast": money.get(slug, {}).get("estimated_monthly_revenue", ""),
                "next_action": decision.get("action", ""),
                "waiting_reason": "User approval" if row.get("approved") == "Pending" else "",
                "blocked_reason": "",
                "missed_deadline": missed,
            }
        )
        rows.append(enriched)
    return rows


def build_ceo_dashboard_rows() -> list[dict[str, Any]]:
    money = build_money_score()
    roi = build_roi_analysis()
    seo = build_seo_difficulty()
    momentum = build_advanced_momentum()
    execution = build_execution_phase2()
    decisions = build_ai_decisions()
    total_revenue = round(sum(numeric(row.get("estimated_monthly_revenue")) for row in money), 2)
    top_money = money[0] if money else {}
    top_roi = roi[0] if roi else {}
    best_seo = sorted(seo, key=lambda row: numeric(row.get("ranking_opportunity")), reverse=True)[0] if seo else {}
    fastest = sorted(momentum, key=lambda row: numeric(row.get("growth_rate")), reverse=True)[0] if momentum else {}
    completion = round(sum(numeric(row.get("completion_percent")) for row in execution) / max(len(execution), 1), 1)
    return [
        {"metric": "Today's Revenue Opportunity", "value": top_money.get("topic", ""), "notes": f"${top_money.get('estimated_monthly_revenue', 0)}/month"},
        {"metric": "Today's Highest ROI Topic", "value": top_roi.get("topic", ""), "notes": f"ROI/hour ${top_roi.get('roi_per_hour', 0)}"},
        {"metric": "Today's Best SEO Opportunity", "value": best_seo.get("topic", ""), "notes": best_seo.get("suggested_angle", "")},
        {"metric": "Today's Fastest Rising Trend", "value": fastest.get("topic", ""), "notes": f"Growth {fastest.get('growth_rate', 0)}%"},
        {"metric": "Topics To Publish Today", "value": len([row for row in decisions if row.get("action") in {"WRITE NOW", "CREATE COMPARISON PAGE", "CREATE LANDING PAGE"}]), "notes": "Requires user approval before publishing"},
        {"metric": "Topics To Refresh Today", "value": len([row for row in decisions if row.get("action") == "REFRESH ARTICLE"]), "notes": ""},
        {"metric": "Topics To Ignore", "value": len([row for row in decisions if row.get("action") in {"IGNORE", "DELETE"}]), "notes": ""},
        {"metric": "Competitor Alerts", "value": len(build_competitor_phase2()), "notes": "Offline/manual data when no API exists"},
        {"metric": "Affiliate Opportunities", "value": len([row for row in build_affiliate_intelligence() if row.get("affiliate_priority") == "High"]), "notes": ""},
        {"metric": "Pending Index Requests", "value": len([row for row in execution if row.get("indexed_google") != "Yes"]), "notes": ""},
        {"metric": "Pending Videos", "value": len([row for row in execution if row.get("youtube_uploaded") != "Yes"]), "notes": ""},
        {"metric": "Pending Social Posts", "value": len([row for row in execution if row.get("facebook_posted") != "Yes" or row.get("linkedin_posted") != "Yes"]), "notes": ""},
        {"metric": "Pending Internal Links", "value": len(read_csv(DATA_DIR / "internal_link_plan.csv")), "notes": ""},
        {"metric": "Execution Progress %", "value": completion, "notes": ""},
        {"metric": "Articles Published Today", "value": 0, "notes": "No publish automation in this dashboard"},
        {"metric": "Videos Published Today", "value": 0, "notes": "No YouTube upload automation"},
        {"metric": "Estimated Monthly Revenue", "value": total_revenue, "notes": ""},
        {"metric": "Estimated Annual Revenue", "value": round(total_revenue * 12, 2), "notes": ""},
        {"metric": "Total Active Opportunities", "value": len(money), "notes": ""},
    ]


def build_ai_summary() -> list[dict[str, str]]:
    roi = build_roi_analysis()
    decisions = build_ai_decisions()
    seo = seo_by_slug()
    if not roi:
        return [{"section": "Executive Summary", "summary": "No ROI data available yet."}]
    best = roi[0]
    decision = next((row for row in decisions if row.get("slug") == best.get("slug")), {})
    seo_row = seo.get(best.get("slug"), {})
    summary = (
        f"Today's best opportunity is {best.get('topic')}. Expected revenue is ${best.get('expected_revenue')}/month. "
        f"Competition is {seo_row.get('competition_level', 'Unknown')} with SEO difficulty {seo_row.get('keyword_difficulty_estimate', 'Unknown')}. "
        f"Recommended action: {decision.get('action', 'WAIT')}. Estimated completion is "
        f"{numeric(best.get('estimated_writing_time')) + numeric(best.get('estimated_video_production_time')) + numeric(best.get('estimated_social_time'))} hours. "
        f"Estimated ROI is ${best.get('roi_per_hour')}/hour."
    )
    return [
        {"section": "Executive Summary", "summary": summary},
        {"section": "Next Step", "summary": "Do not auto-publish. Review the top P1/P2 items, approve one article plan, then run the separate website publishing workflow only when ready."},
    ]


def write_phase2_outputs() -> dict[str, Any]:
    ceo = build_ceo_dashboard_rows()
    roi = build_roi_analysis()
    decisions = build_ai_decisions()
    predictions = build_predictions()
    momentum = build_advanced_momentum()
    calendar = build_content_calendar()
    competitor = build_competitor_phase2()
    gap = build_content_gap()
    affiliate = build_affiliate_intelligence()
    pipeline = build_article_pipeline()
    execution = build_execution_phase2()
    summary = build_ai_summary()

    outputs = [
        ("ceo_dashboard", ceo, CEO_FIELDS),
        ("roi_analysis", roi, ROI_FIELDS),
        ("ai_decision_engine", decisions, DECISION_FIELDS),
        ("predictions", predictions, PREDICTION_FIELDS),
        ("advanced_trend_momentum", momentum, ADVANCED_MOMENTUM_FIELDS),
        ("content_calendar", calendar, CALENDAR_FIELDS),
        ("competitor_phase2", competitor, COMPETITOR_PHASE2_FIELDS),
        ("content_gap", gap, CONTENT_GAP_FIELDS),
        ("affiliate_intelligence", affiliate, AFFILIATE_INTEL_FIELDS),
        ("article_pipeline", pipeline, ARTICLE_PIPELINE_FIELDS),
        ("execution_tracker_phase2", execution, EXECUTION_PHASE2_FIELDS),
        ("ai_executive_summary", summary, AI_SUMMARY_FIELDS),
    ]
    for name, rows, fields in outputs:
        write_csv(DATA_DIR / f"{name}.csv", rows, fields)
        write_json(DATA_DIR / f"{name}.json", rows)
    update_master_workbook(
        {
            "CEO Dashboard": (ceo, CEO_FIELDS),
            "ROI Analysis": (roi, ROI_FIELDS),
            "AI Decision Engine": (decisions, DECISION_FIELDS),
            "Predictions": (predictions, PREDICTION_FIELDS),
            "Advanced Momentum": (momentum, ADVANCED_MOMENTUM_FIELDS),
            "Content Calendar": (calendar, CALENDAR_FIELDS),
            "Competitor Intelligence": (competitor, COMPETITOR_PHASE2_FIELDS),
            "Content Gap": (gap, CONTENT_GAP_FIELDS),
            "Affiliate Intelligence": (affiliate, AFFILIATE_INTEL_FIELDS),
            "Article Pipeline": (pipeline, ARTICLE_PIPELINE_FIELDS),
            "Execution Tracker+": (execution, EXECUTION_PHASE2_FIELDS),
            "AI Executive Summary": (summary, AI_SUMMARY_FIELDS),
        }
    )
    apply_excel_polish(DATA_DIR / "master_dashboard.xlsx")
    return {name: len(rows) for name, rows, _fields in outputs}


def apply_excel_polish(path: Path) -> bool:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False
    if not path.exists():
        return False
    workbook = load_workbook(path)
    if "CEO Dashboard" in workbook.sheetnames:
        ws = workbook["CEO Dashboard"]
        workbook._sheets.remove(ws)
        workbook._sheets.insert(0, ws)
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for row in ws.iter_rows(min_row=2):
            text = " ".join(str(cell.value or "") for cell in row)
            fill = None
            if "P1" in text or "WRITE NOW" in text:
                fill = PatternFill("solid", fgColor="D9EAD3")
            elif "P2" in text or "WAIT" in text:
                fill = PatternFill("solid", fgColor="FFF2CC")
            elif "IGNORE" in text or "DELETE" in text:
                fill = PatternFill("solid", fgColor="F4CCCC")
            if fill:
                for cell in row:
                    cell.fill = fill
        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 60)
    add_charts_sheet(workbook, Font)
    result = safe_save_workbook(workbook, path)
    print_workbook_report(result)
    return bool(result.get("saved") or result.get("pending"))


def add_charts_sheet(workbook: Any, Font: Any) -> None:
    if "Charts" in workbook.sheetnames:
        del workbook["Charts"]
    ws = workbook.create_sheet("Charts")
    ws["A1"] = "Dashboard Charts"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Excel-safe mode is enabled. Chart objects are intentionally disabled to avoid workbook recovery warnings."
    ws["A5"] = "Chart themes tracked for future safe rendering:"
    ws["A6"] = "Revenue Trend, Opportunity Trend, SEO Difficulty, Competitor Comparison, Execution Status, Momentum Distribution, Affiliate Revenue, Content Status, Publish Frequency, Monthly Growth."
