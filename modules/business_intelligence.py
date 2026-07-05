from __future__ import annotations

import html
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from modules.opportunity_forecast import build_forecast
from modules.performance_tracking import BASE_URL, DATA_DIR, load_json_rows, numeric, print_workbook_report, read_csv, safe_save_workbook, slug_from_url, slugify, update_master_workbook, write_csv, write_json


OPPORTUNITY_FIELDS = [
    "rank",
    "topic",
    "slug",
    "intent",
    "seo_score",
    "traffic_score",
    "buyer_intent_score",
    "revenue_score",
    "competition_score",
    "freshness_score",
    "video_score",
    "social_score",
    "total_score",
    "reason",
]

MONEY_FIELDS = [
    "rank",
    "topic",
    "slug",
    "estimated_monthly_traffic",
    "estimated_ctr",
    "estimated_affiliate_clicks",
    "estimated_conversion_rate",
    "estimated_commission",
    "estimated_monthly_revenue",
    "revenue_confidence",
    "recommended_affiliate_program",
    "money_score",
]

SEO_DIFFICULTY_FIELDS = [
    "topic",
    "slug",
    "keyword_difficulty_estimate",
    "competition_level",
    "serp_risk",
    "ranking_opportunity",
    "suggested_angle",
    "recommendation_adjustment",
]

ARTICLE_PLAN_FIELDS = [
    "priority",
    "suggested_title",
    "slug",
    "meta_description",
    "article_type",
    "primary_keyword",
    "secondary_keywords",
    "recommended_outline",
    "cta_idea",
    "affiliate_angle",
    "internal_links_needed",
    "youtube_needed",
    "social_needed",
]

EXECUTION_FIELDS = [
    "topic",
    "slug",
    "article_url",
    "collected",
    "approved",
    "article_created",
    "website_published",
    "youtube_script_created",
    "youtube_uploaded",
    "facebook_posted",
    "linkedin_posted",
    "quora_answered",
    "producthunt_commented",
    "indexed_google",
    "indexed_bing",
    "indexed_yandex",
    "revenue_checked",
]

TREND_MOMENTUM_FIELDS = [
    "topic",
    "slug",
    "first_seen_date",
    "last_seen_date",
    "times_seen",
    "rank_today",
    "rank_yesterday",
    "rank_change",
    "score_today",
    "score_7d_avg",
    "score_30d_avg",
    "momentum",
    "trend_age_days",
]

AI_RECOMMENDATION_FIELDS = [
    "rank",
    "topic",
    "slug",
    "final_action",
    "reason",
    "expected_impact",
    "risk",
    "next_step",
]

COMPETITOR_INTEL_FIELDS = [
    "topic",
    "competitor_url",
    "competitor_title",
    "estimated_strength",
    "content_gap",
    "backlink_gap",
    "freshness_gap",
    "our_angle",
]

AFFILIATE_MATCH_FIELDS = [
    "topic",
    "slug",
    "PartnerStack",
    "Impact",
    "CJ",
    "ShareASale",
    "Direct affiliate",
    "No clear program",
    "program_confidence",
    "commission_estimate",
    "affiliate_priority",
    "notes",
]

DASHBOARD_FIELDS = ["section", "rank", "topic", "slug", "score", "action", "reason"]

COMMERCIAL_TERMS = ("review", "comparison", "best", "alternative", "alternatives", "pricing", "coupon", "discount", "tutorial", "guide", " vs ")


def topic_rows() -> list[dict[str, Any]]:
    rows = load_json_rows(DATA_DIR / "topic_scores.json")
    if rows:
        return rows
    return load_json_rows(DATA_DIR / "topic_demo_scores.json")


def score_average(score_map: dict[str, Any]) -> float:
    values = [numeric(value) for value in score_map.values()]
    return round(sum(values) / max(len(values), 1), 1)


def infer_intent(topic: str) -> str:
    text = topic.lower()
    if " vs " in text or "-vs-" in text or "comparison" in text or "compare" in text:
        return "comparison"
    if "alternatives" in text or "alternative" in text:
        return "alternative"
    if "pricing" in text or "cost" in text:
        return "pricing"
    if "best" in text or "top " in text:
        return "best list"
    if "tutorial" in text or "how to" in text or "guide" in text:
        return "tutorial"
    if "review" in text:
        return "review"
    if "news" in text or "funding" in text or "launch" in text:
        return "news"
    if "reddit" in text or "discussion" in text:
        return "social discussion"
    return "evergreen"


def build_opportunity_breakdown(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    source_rows = rows if rows is not None else topic_rows()
    output = []
    for index, row in enumerate(source_rows, 1):
        topic = str(row.get("topic", ""))
        competition = numeric(row.get("competition") or row.get("competition_level"))
        buyer = numeric(row.get("buyer_intent") or row.get("buyer_intent_score"))
        revenue = numeric(row.get("revenue_score") or row.get("affiliate_value"))
        seo = numeric(row.get("seo_score") or row.get("seo_opportunity"))
        traffic = numeric(row.get("traffic_score") or row.get("estimated_traffic"))
        freshness = numeric(row.get("freshness") or row.get("trend_score"))
        video = numeric(row.get("video_score") or row.get("youtube_potential"))
        social = score_average(row.get("social_scores", {})) if isinstance(row.get("social_scores"), dict) else numeric(row.get("social_score"))
        competition_score = max(0, 100 - competition)
        total = numeric(row.get("total_score")) or round(
            seo * 0.16 + traffic * 0.14 + buyer * 0.18 + revenue * 0.18 + competition_score * 0.11 + freshness * 0.08 + video * 0.08 + social * 0.07,
            1,
        )
        output.append(
            {
                "rank": index,
                "topic": topic,
                "slug": slugify(row.get("slug") or topic),
                "intent": infer_intent(topic),
                "seo_score": round(seo, 1),
                "traffic_score": round(traffic, 1),
                "buyer_intent_score": round(buyer, 1),
                "revenue_score": round(revenue, 1),
                "competition_score": round(competition_score, 1),
                "freshness_score": round(freshness, 1),
                "video_score": round(video, 1),
                "social_score": round(social, 1),
                "total_score": round(total, 1),
                "reason": explain_score(topic, total, buyer, revenue, competition, freshness),
            }
        )
    return sorted(output, key=lambda item: numeric(item.get("total_score")), reverse=True)


def explain_score(topic: str, total: float, buyer: float, revenue: float, competition: float, freshness: float) -> str:
    reasons = []
    if any(term in topic.lower() for term in COMMERCIAL_TERMS):
        reasons.append("commercial keyword")
    if buyer >= 70:
        reasons.append("high buyer intent")
    if revenue >= 65:
        reasons.append("strong affiliate value")
    if competition >= 70:
        reasons.append("high competition risk")
    if freshness >= 70:
        reasons.append("fresh trend signal")
    if total < 60:
        reasons.append("overall score is still weak")
    return "; ".join(reasons) or "balanced opportunity"


def estimate_ctr(intent: str, ranking_opportunity: float) -> float:
    base = {"review": 0.035, "comparison": 0.032, "pricing": 0.04, "best list": 0.028, "alternative": 0.03}.get(intent, 0.018)
    return round(base * (0.7 + ranking_opportunity / 160), 4)


def affiliate_program_for(topic: str) -> tuple[str, float, str]:
    text = topic.lower()
    if any(term in text for term in ("saas", "automation", "email", "crm", "affiliate", "marketing")):
        return "PartnerStack / Direct affiliate", 75, "High"
    if any(term in text for term in ("website", "hosting", "builder")):
        return "Impact / Direct affiliate", 70, "High"
    if any(term in text for term in ("seo", "keyword", "content optimization")):
        return "Direct affiliate / PartnerStack", 68, "High"
    if any(term in text for term in ("ai", "software", "tool")):
        return "Direct affiliate", 55, "Medium"
    return "No clear program", 25, "Low"


def build_money_score(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    breakdown = build_opportunity_breakdown(rows)
    forecast_by_slug = {row.get("slug"): row for row in build_forecast().rows}
    output = []
    for index, row in enumerate(breakdown, 1):
        forecast = forecast_by_slug.get(row.get("slug"), {})
        intent = row.get("intent", "")
        ranking = max(0, 100 - numeric(forecast.get("estimated_difficulty")))
        traffic = numeric(forecast.get("estimated_monthly_traffic"), numeric(row.get("traffic_score")) * 18)
        ctr = estimate_ctr(intent, ranking)
        affiliate_clicks = round(traffic * ctr, 1)
        conversion_rate = round(0.018 + numeric(row.get("buyer_intent_score")) / 4000 + numeric(row.get("revenue_score")) / 5000, 4)
        program, confidence, priority = affiliate_program_for(str(row.get("topic")))
        commission = round(8 + numeric(row.get("revenue_score")) * 0.28, 2)
        revenue = round(affiliate_clicks * conversion_rate * commission, 2)
        output.append(
            {
                "rank": index,
                "topic": row.get("topic", ""),
                "slug": row.get("slug", ""),
                "estimated_monthly_traffic": int(traffic),
                "estimated_ctr": ctr,
                "estimated_affiliate_clicks": affiliate_clicks,
                "estimated_conversion_rate": conversion_rate,
                "estimated_commission": commission,
                "estimated_monthly_revenue": revenue,
                "revenue_confidence": confidence,
                "recommended_affiliate_program": program,
                "money_score": round(numeric(row.get("buyer_intent_score")) * 0.28 + numeric(row.get("revenue_score")) * 0.32 + ranking * 0.15 + confidence * 0.25, 1),
            }
        )
    return sorted(output, key=lambda item: (numeric(item.get("money_score")), numeric(item.get("estimated_monthly_revenue"))), reverse=True)


def build_seo_difficulty(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    breakdown = build_opportunity_breakdown(rows)
    output = []
    for row in breakdown:
        difficulty = max(0, 100 - numeric(row.get("competition_score")))
        level = "Low" if difficulty < 40 else "Medium" if difficulty < 70 else "High"
        buyer = numeric(row.get("buyer_intent_score"))
        adjustment = "Refresh / Long-tail / Support article" if buyer >= 65 and level == "High" else "Primary article" if buyer >= 60 else "Monitor"
        output.append(
            {
                "topic": row.get("topic", ""),
                "slug": row.get("slug", ""),
                "keyword_difficulty_estimate": round(difficulty, 1),
                "competition_level": level,
                "serp_risk": "High SERP authority risk" if level == "High" else "Moderate SERP risk" if level == "Medium" else "Lower SERP risk",
                "ranking_opportunity": round(max(0, 100 - difficulty) * 0.55 + numeric(row.get("seo_score")) * 0.45, 1),
                "suggested_angle": suggested_angle(str(row.get("topic")), str(row.get("intent"))),
                "recommendation_adjustment": adjustment,
            }
        )
    return output


def suggested_angle(topic: str, intent: str) -> str:
    if intent == "comparison":
        return "Buyer-focused comparison with pricing, fit, and alternatives"
    if intent == "pricing":
        return "Pricing explainer with official-price verification warning"
    if intent in {"review", "alternative", "best list"}:
        return "Practical review angle with use cases and risks"
    return f"Long-tail support article for {topic}"


def build_article_plan(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    recommendations = build_ai_recommendations(rows)
    output = []
    for rec in recommendations:
        if rec.get("final_action") not in {"Write Now", "Write Today", "Make Video"}:
            continue
        topic = str(rec.get("topic"))
        slug = str(rec.get("slug"))
        intent = infer_intent(topic)
        priority = "P1" if rec.get("final_action") == "Write Now" else "P2"
        output.append(
            {
                "priority": priority,
                "suggested_title": title_for_topic(topic, intent),
                "slug": slug,
                "meta_description": meta_for_topic(topic),
                "article_type": intent,
                "primary_keyword": topic.lower(),
                "secondary_keywords": secondary_keywords(topic),
                "recommended_outline": outline_for_intent(intent),
                "cta_idea": "Read the full review and verify current pricing on the official website.",
                "affiliate_angle": affiliate_program_for(topic)[0],
                "internal_links_needed": "Add 5-10 related review, comparison, pricing, and category links.",
                "youtube_needed": "yes" if rec.get("final_action") in {"Write Now", "Make Video"} else "no",
                "social_needed": "yes" if rec.get("final_action") in {"Write Now", "Write Today"} else "no",
            }
        )
    return output[:30]


def title_for_topic(topic: str, intent: str) -> str:
    if "2026" in topic:
        return topic
    suffix = "Review 2026" if intent == "review" else "Guide 2026"
    return f"{topic}: {suffix}"


def meta_for_topic(topic: str) -> str:
    return f"Compare {topic} by pricing, features, pros, cons, alternatives, buyer fit, and practical risks before choosing a tool."


def secondary_keywords(topic: str) -> str:
    base = slugify(topic).replace("-", " ")
    return ", ".join([f"{base} pricing", f"{base} alternatives", f"{base} review", f"{base} pros and cons"])


def outline_for_intent(intent: str) -> str:
    common = ["Overview", "Best for / Not best for", "Pricing", "Pros and cons", "Alternatives", "FAQ", "Final verdict"]
    if intent == "comparison":
        common.insert(1, "Feature comparison table")
    if intent == "pricing":
        common.insert(1, "Plan and cost breakdown")
    return " | ".join(common)


def load_upload_status() -> dict[str, dict[str, str]]:
    rows = []
    for path in (DATA_DIR / "upload_links.csv", Path("video_output") / "upload_links.csv"):
        rows.extend(read_csv(path))
    result = {}
    for row in rows:
        slug = slugify(row.get("FolderName") or slug_from_url(row.get("PageUrl", "")))
        if slug:
            result[slug] = row
    return result


def build_execution_tracker(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    breakdown = build_opportunity_breakdown(rows)
    uploads = load_upload_status()
    lifecycle = {row.get("slug"): row for row in read_csv(DATA_DIR / "content_lifecycle.csv")}
    output = []
    for row in breakdown:
        slug = str(row.get("slug"))
        upload = uploads.get(slug, {})
        life = lifecycle.get(slug, {})
        article_url = upload.get("PageUrl") or life.get("article_url") or f"{BASE_URL}/{slug}/"
        has_video = bool(upload.get("YoutubeVideoUrl") or life.get("youtube_url"))
        output.append(
            {
                "topic": row.get("topic", ""),
                "slug": slug,
                "article_url": article_url,
                "collected": "Yes",
                "approved": "Pending",
                "article_created": "Yes" if life.get("article_url") else "No",
                "website_published": "Yes" if life.get("article_url") else "No",
                "youtube_script_created": "Pending",
                "youtube_uploaded": "Yes" if has_video else "No",
                "facebook_posted": "No",
                "linkedin_posted": "No",
                "quora_answered": "No",
                "producthunt_commented": "No",
                "indexed_google": "Pending",
                "indexed_bing": "Pending",
                "indexed_yandex": "Pending",
                "revenue_checked": "No",
            }
        )
    return output


def build_trend_momentum(now: datetime | None = None) -> list[dict[str, Any]]:
    current_date = (now or datetime.now(timezone.utc)).date()
    history = read_csv(DATA_DIR / "hottrend_topic_history.csv")
    if not history:
        history = [
            {
                "run_date": current_date.isoformat(),
                "topic": row.get("topic", ""),
                "slug": row.get("slug", ""),
                "rank_today": row.get("rank", ""),
                "total_score": row.get("total_score", ""),
            }
            for row in build_opportunity_breakdown()
        ]
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in history:
        slug = slugify(row.get("slug") or row.get("topic"))
        if slug:
            grouped[slug].append(row)
    output = []
    for slug, rows in grouped.items():
        rows = sorted(rows, key=lambda row: row.get("run_date", ""))
        latest = rows[-1]
        previous = rows[-2] if len(rows) > 1 else {}
        first_date = rows[0].get("first_seen_date") or rows[0].get("run_date") or current_date.isoformat()
        last_date = latest.get("last_seen_date") or latest.get("run_date") or current_date.isoformat()
        score_today = numeric(latest.get("total_score") or latest.get("score_today"))
        scores = [numeric(row.get("total_score") or row.get("score_today")) for row in rows]
        rank_today = int(numeric(latest.get("rank_today") or latest.get("rank")))
        rank_yesterday = int(numeric(previous.get("rank_today") or previous.get("rank"))) if previous else 0
        rank_change = rank_yesterday - rank_today if rank_yesterday and rank_today else 0
        score_7d = round(sum(scores[-7:]) / max(len(scores[-7:]), 1), 1)
        score_30d = round(sum(scores[-30:]) / max(len(scores[-30:]), 1), 1)
        momentum = "Rising" if score_today > score_7d and rank_change > 0 else "Declining" if score_today < score_7d and rank_change < 0 else "Stable"
        output.append(
            {
                "topic": latest.get("topic", ""),
                "slug": slug,
                "first_seen_date": first_date,
                "last_seen_date": last_date,
                "times_seen": len(rows),
                "rank_today": rank_today,
                "rank_yesterday": rank_yesterday,
                "rank_change": rank_change,
                "score_today": score_today,
                "score_7d_avg": score_7d,
                "score_30d_avg": score_30d,
                "momentum": momentum,
                "trend_age_days": max(0, (current_date - parse_date(first_date, current_date)).days),
            }
        )
    return sorted(output, key=lambda row: (numeric(row.get("score_today")), numeric(row.get("rank_change"))), reverse=True)


def parse_date(value: Any, fallback) -> Any:
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except ValueError:
        return fallback


def build_competitor_intelligence() -> list[dict[str, Any]]:
    watch = read_csv(DATA_DIR / "competitor_watch.csv")
    if not watch:
        return [
            {
                "topic": row.get("topic", ""),
                "competitor_url": "Needs manual input",
                "competitor_title": "Needs manual input",
                "estimated_strength": "Needs manual input",
                "content_gap": "Needs manual input",
                "backlink_gap": "Needs manual input",
                "freshness_gap": "Needs manual input",
                "our_angle": suggested_angle(row.get("topic", ""), infer_intent(row.get("topic", ""))),
            }
            for row in build_opportunity_breakdown()[:20]
        ]
    return [
        {
            "topic": row.get("target_keyword") or row.get("new_opportunity", ""),
            "competitor_url": row.get("competitor_url", "") or "Needs manual input",
            "competitor_title": row.get("page_type", "") or "Needs manual input",
            "estimated_strength": row.get("difficulty", "") or "Needs manual input",
            "content_gap": row.get("new_opportunity", "") or "Needs manual input",
            "backlink_gap": "Needs manual input",
            "freshness_gap": row.get("publishing_frequency", "") or "Needs manual input",
            "our_angle": row.get("suggested_response", "") or "Create a stronger buyer-focused page",
        }
        for row in watch
    ]


def build_affiliate_match(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    breakdown = build_opportunity_breakdown(rows)
    output = []
    for row in breakdown:
        topic = str(row.get("topic"))
        program, confidence, priority = affiliate_program_for(topic)
        output.append(
            {
                "topic": topic,
                "slug": row.get("slug", ""),
                "PartnerStack": "Possible" if "PartnerStack" in program else "Check",
                "Impact": "Possible" if "Impact" in program else "Check",
                "CJ": "Check",
                "ShareASale": "Check",
                "Direct affiliate": "Likely" if "Direct" in program else "Check",
                "No clear program": "Yes" if program == "No clear program" else "No",
                "program_confidence": confidence,
                "commission_estimate": round(8 + numeric(row.get("revenue_score")) * 0.28, 2),
                "affiliate_priority": priority,
                "notes": f"Best initial match: {program}. Verify terms before publishing affiliate CTA.",
            }
        )
    return output


def build_ai_recommendations(rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    breakdown = build_opportunity_breakdown(rows)
    money = {row.get("slug"): row for row in build_money_score(rows)}
    seo = {row.get("slug"): row for row in build_seo_difficulty(rows)}
    output = []
    for index, row in enumerate(breakdown, 1):
        slug = row.get("slug")
        money_row = money.get(slug, {})
        seo_row = seo.get(slug, {})
        action = final_action(row, money_row, seo_row)
        risk = "High" if seo_row.get("competition_level") == "High" and numeric(money_row.get("revenue_confidence")) < 65 else "Medium" if seo_row.get("competition_level") == "High" else "Low"
        output.append(
            {
                "rank": index,
                "topic": row.get("topic", ""),
                "slug": slug,
                "final_action": action,
                "reason": row.get("reason", ""),
                "expected_impact": expected_impact(action, money_row),
                "risk": risk,
                "next_step": next_step(action, index),
            }
        )
    return output


def final_action(row: dict[str, Any], money_row: dict[str, Any], seo_row: dict[str, Any]) -> str:
    total = numeric(row.get("total_score"))
    money_score = numeric(money_row.get("money_score"))
    video = numeric(row.get("video_score"))
    if total >= 82 and money_score >= 70:
        return "Write Now"
    if total >= 72 and money_score >= 60:
        return "Write Today"
    if video >= 70 and money_score >= 55:
        return "Make Video"
    if seo_row.get("recommendation_adjustment") == "Refresh / Long-tail / Support article":
        return "Refresh Existing"
    if total >= 58:
        return "Watch"
    return "Skip"


def expected_impact(action: str, money_row: dict[str, Any]) -> str:
    traffic = money_row.get("estimated_monthly_traffic", 0)
    revenue = money_row.get("estimated_monthly_revenue", 0)
    if action in {"Write Now", "Write Today"}:
        return f"Potential monthly traffic {traffic}; estimated revenue ${revenue}"
    if action == "Make Video":
        return "Support YouTube growth and improve article engagement"
    if action == "Refresh Existing":
        return "Improve CTR, rankings, and internal conversion"
    return "Low immediate impact"


def next_step(action: str, rank: int) -> str:
    if rank <= 3 and action in {"Write Now", "Write Today"}:
        return "Write this article"
    if rank <= 3 and action == "Make Video":
        return "Create YouTube script"
    if rank <= 3:
        return "Post on social manually"
    return {
        "Write Now": "Prepare article brief for approval",
        "Write Today": "Add to today's writing queue",
        "Make Video": "Create YouTube script",
        "Refresh Existing": "Refresh existing page",
        "Watch": "Monitor for one more run",
        "Skip": "Do not write now",
    }.get(action, "Monitor")


def build_dashboard_rows() -> list[dict[str, Any]]:
    recommendations = build_ai_recommendations()
    money = build_money_score()
    seo = build_seo_difficulty()
    rows: list[dict[str, Any]] = []
    sections = [
        ("Today's Top 10", recommendations[:10], "rank", "final_action"),
        ("Top 3 Write Now", [row for row in recommendations if row.get("final_action") in {"Write Now", "Write Today"}][:3], "rank", "final_action"),
        ("Top 3 Money Opportunities", money[:3], "rank", "money_score"),
        ("Top 3 SEO Easy Wins", sorted(seo, key=lambda row: numeric(row.get("ranking_opportunity")), reverse=True)[:3], "ranking_opportunity", "competition_level"),
        ("Top 3 Video Candidates", [row for row in build_opportunity_breakdown() if numeric(row.get("video_score")) >= 60][:3], "rank", "video_score"),
        ("Top 3 Refresh Candidates", [row for row in recommendations if row.get("final_action") == "Refresh Existing"][:3], "rank", "final_action"),
        ("Warning / Risky Topics", [row for row in recommendations if row.get("risk") == "High" or row.get("final_action") == "Skip"][:10], "rank", "risk"),
    ]
    for section, section_rows, rank_field, action_field in sections:
        for idx, row in enumerate(section_rows, 1):
            rows.append(
                {
                    "section": section,
                    "rank": row.get("rank") or idx,
                    "topic": row.get("topic", ""),
                    "slug": row.get("slug", ""),
                    "score": row.get("total_score") or row.get("money_score") or row.get("ranking_opportunity") or "",
                    "action": row.get("final_action") or row.get("recommended_action") or row.get(action_field, ""),
                    "reason": row.get("reason") or row.get("suggested_angle") or "",
                }
            )
    return rows


def write_daily_ceo_dashboard_html(path: Path, dashboard_rows: list[dict[str, Any]], recommendations: list[dict[str, Any]]) -> None:
    def table(rows: list[dict[str, Any]], fields: list[str]) -> str:
        head = "".join(f"<th>{html.escape(field.replace('_', ' ').title())}</th>" for field in fields)
        body = "".join("<tr>" + "".join(f"<td>{html.escape(str(row.get(field, '')))}</td>" for field in fields) + "</tr>" for row in rows)
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    doc = f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>AI CEO Business Intelligence Dashboard</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;background:#f8fafc;color:#172033}}table{{width:100%;border-collapse:collapse;background:#fff;margin:12px 0 28px}}td,th{{border:1px solid #d8e1ea;padding:8px;text-align:left}}th{{background:#e6fffb}}.p1{{color:#0f766e;font-weight:700}}.risk{{color:#991b1b;font-weight:700}}</style></head>
<body><h1>AI CEO Business Intelligence Dashboard</h1>
<p>Recommendation-only dashboard. No article publishing, deployment, YouTube upload, or social posting.</p>
<h2>Dashboard</h2>{table(dashboard_rows, DASHBOARD_FIELDS)}
<h2>AI Recommendation</h2>{table(recommendations, AI_RECOMMENDATION_FIELDS)}
</body></html>"""
    path.write_text(doc, encoding="utf-8")


def apply_dashboard_formatting(path: Path) -> bool:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        return False
    if not path.exists():
        return False
    workbook = load_workbook(path)
    for sheet_name in ("Dashboard", "AI Recommendation", "Money Score"):
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            text = " ".join(str(cell.value or "") for cell in row)
            fill = None
            if "Write Now" in text or "P1" in text:
                fill = PatternFill("solid", fgColor="D9EAD3")
            elif "Write Today" in text or "P2" in text or "Watch" in text:
                fill = PatternFill("solid", fgColor="FFF2CC")
            elif "Skip" in text:
                fill = PatternFill("solid", fgColor="F4CCCC")
            if fill:
                for cell in row:
                    cell.fill = fill
    result = safe_save_workbook(workbook, path)
    print_workbook_report(result)
    return bool(result.get("saved") or result.get("pending"))


def write_business_intelligence_outputs() -> dict[str, Any]:
    opportunity = build_opportunity_breakdown()
    money = build_money_score()
    seo = build_seo_difficulty()
    article_plan = build_article_plan()
    execution = build_execution_tracker()
    momentum = build_trend_momentum()
    competitor = build_competitor_intelligence()
    affiliate = build_affiliate_match()
    recommendations = build_ai_recommendations()
    dashboard = build_dashboard_rows()

    write_csv(DATA_DIR / "opportunity_score_breakdown.csv", opportunity, OPPORTUNITY_FIELDS)
    write_json(DATA_DIR / "opportunity_score_breakdown.json", opportunity)
    write_csv(DATA_DIR / "money_score.csv", money, MONEY_FIELDS)
    write_json(DATA_DIR / "money_score.json", money)
    write_csv(DATA_DIR / "seo_difficulty.csv", seo, SEO_DIFFICULTY_FIELDS)
    write_json(DATA_DIR / "seo_difficulty.json", seo)
    write_csv(DATA_DIR / "article_plan.csv", article_plan, ARTICLE_PLAN_FIELDS)
    write_json(DATA_DIR / "article_plan.json", article_plan)
    write_csv(DATA_DIR / "execution_tracker.csv", execution, EXECUTION_FIELDS)
    write_json(DATA_DIR / "execution_tracker.json", execution)
    write_csv(DATA_DIR / "trend_momentum.csv", momentum, TREND_MOMENTUM_FIELDS)
    write_json(DATA_DIR / "trend_momentum.json", momentum)
    write_csv(DATA_DIR / "competitor_intelligence.csv", competitor, COMPETITOR_INTEL_FIELDS)
    write_json(DATA_DIR / "competitor_intelligence.json", competitor)
    write_csv(DATA_DIR / "affiliate_match.csv", affiliate, AFFILIATE_MATCH_FIELDS)
    write_json(DATA_DIR / "affiliate_match.json", affiliate)
    write_csv(DATA_DIR / "ai_recommendations.csv", recommendations, AI_RECOMMENDATION_FIELDS)
    write_json(DATA_DIR / "ai_recommendations.json", recommendations)
    write_daily_ceo_dashboard_html(DATA_DIR / "daily_ceo_dashboard.html", dashboard, recommendations)
    update_master_workbook(
        {
            "Dashboard": (dashboard, DASHBOARD_FIELDS),
            "Opportunity Score": (opportunity, OPPORTUNITY_FIELDS),
            "Money Score": (money, MONEY_FIELDS),
            "SEO Difficulty": (seo, SEO_DIFFICULTY_FIELDS),
            "Article Plan": (article_plan, ARTICLE_PLAN_FIELDS),
            "Execution Tracker": (execution, EXECUTION_FIELDS),
            "Trend Momentum": (momentum, TREND_MOMENTUM_FIELDS),
            "AI Recommendation": (recommendations, AI_RECOMMENDATION_FIELDS),
            "Competitor Watch": (competitor, COMPETITOR_INTEL_FIELDS),
            "Affiliate Match": (affiliate, AFFILIATE_MATCH_FIELDS),
        }
    )
    formatted = apply_dashboard_formatting(DATA_DIR / "master_dashboard.xlsx")
    return {
        "opportunity": len(opportunity),
        "money": len(money),
        "seo": len(seo),
        "article_plan": len(article_plan),
        "execution": len(execution),
        "momentum": len(momentum),
        "competitor": len(competitor),
        "affiliate": len(affiliate),
        "recommendations": len(recommendations),
        "formatted": formatted,
    }
