from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from modules.content_operations import (
    AI_PRIORITY_DASHBOARD_FIELDS,
    CONTENT_GAP_FIELDS,
    REVENUE_OPPORTUNITY_FIELDS,
    TODAY_WRITE_PLAN_FIELDS,
    WEBSITE_PUBLISHING_QUEUE_FIELDS,
)
from modules.daily_content_factory import (
    INTERNAL_LINK_INSERTION_FIELDS,
    REFRESH_QUEUE_FIELDS,
    TODAY_SELECTED_TOPICS_FIELDS,
)
from modules.performance_tracking import (
    DATA_DIR,
    clean_excel_value,
    clean_sheet_name,
    numeric,
    read_csv,
    safe_save_workbook,
    write_csv,
    write_json,
)


MORNING_COMMAND_FIELDS = [
    "rank",
    "topic",
    "slug",
    "final_score",
    "priority",
    "action",
    "article_type",
    "target_keyword",
    "duplicate_risk",
    "estimated_value",
    "publish_status",
    "published_url",
    "video_package_path",
    "next_action",
    "note",
]

PUBLISHED_TODAY_FIELDS = [
    "run_timestamp",
    "topic",
    "slug",
    "article_type",
    "status",
    "published_url",
    "source_file",
    "word_count",
    "internal_links_added",
    "affiliate_links_added",
    "video_package_path",
    "error",
]

ARTICLE_DRAFT_DAILY_FIELDS = [
    "run_timestamp",
    "topic",
    "slug",
    "article_type",
    "status",
    "markdown_path",
    "json_path",
    "word_count",
    "youtube_embed_position",
    "error",
]

VIDEO_PACKAGE_DAILY_FIELDS = [
    "run_timestamp",
    "topic",
    "slug",
    "article_type",
    "status",
    "video_package_path",
    "files_created",
    "required_files_present",
    "error",
]

FRONT_SHEETS = [
    "Morning Command Center",
    "CEO Dashboard",
    "Today Selected Topics",
    "Today Write Plan",
    "Website Publishing Queue",
    "Article Draft Report",
    "Video Package Report",
    "Published Today",
    "Refresh Queue",
    "CTR Title Tests",
    "AI Priority Dashboard",
    "Revenue Opportunity",
    "Money Ranking",
    "Duplicate Risk",
    "Content Gap",
    "Internal Link Insertions",
    "Top SEO",
    "Top Revenue",
    "Top Video",
    "Top Social",
    "Full History",
]

DEFAULT_SHEET_FIELDS = {
    "Morning Command Center": MORNING_COMMAND_FIELDS,
    "CEO Dashboard": ["section", "value", "note"],
    "Today Selected Topics": TODAY_SELECTED_TOPICS_FIELDS,
    "Today Write Plan": TODAY_WRITE_PLAN_FIELDS,
    "Website Publishing Queue": WEBSITE_PUBLISHING_QUEUE_FIELDS,
    "Article Draft Report": ARTICLE_DRAFT_DAILY_FIELDS,
    "Video Package Report": VIDEO_PACKAGE_DAILY_FIELDS,
    "Published Today": PUBLISHED_TODAY_FIELDS,
    "Refresh Queue": REFRESH_QUEUE_FIELDS,
    "CTR Title Tests": ["slug", "topic", "variant", "title", "intent", "status"],
    "AI Priority Dashboard": AI_PRIORITY_DASHBOARD_FIELDS,
    "Revenue Opportunity": REVENUE_OPPORTUNITY_FIELDS,
    "Money Ranking": [
        "rank",
        "topic",
        "slug",
        "money_score",
        "decision",
        "reason",
    ],
    "Duplicate Risk": [
        "topic",
        "slug",
        "matched_slug",
        "matched_url",
        "duplicate_score",
        "decision",
        "reason",
    ],
    "Content Gap": CONTENT_GAP_FIELDS,
    "Internal Link Insertions": INTERNAL_LINK_INSERTION_FIELDS,
    "Top SEO": ["rank", "topic", "slug", "score", "action"],
    "Top Revenue": ["rank", "topic", "slug", "estimated_value", "action"],
    "Top Video": ["rank", "topic", "slug", "youtube_score", "action"],
    "Top Social": ["rank", "topic", "slug", "social_score", "action"],
    "Full History": ["run_timestamp", "topic", "slug", "status", "note"],
}


def _by_slug(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(row.get("slug", "")): row for row in rows if row.get("slug")}


def _next_action(row: dict[str, Any], publish: dict[str, Any], video: dict[str, Any]) -> str:
    decision = str(row.get("decision") or row.get("action") or "")
    publish_status = str(publish.get("status", ""))
    video_status = str(video.get("status", ""))
    if "error" in publish_status.lower():
        return "Review publish error"
    if decision == "REFRESH_EXISTING":
        return "Review refreshed source and update existing page"
    if publish_status.startswith("published") and video_status:
        return "Upload video to YouTube manually"
    if publish_status == "preview":
        return "Review article preview before publishing"
    if video_status:
        return "Review video package"
    return "Review topic decision"


def build_morning_command_rows(
    selected_rows: list[dict[str, Any]],
    publish_reports: list[dict[str, Any]],
    video_reports: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    publish_by_slug = _by_slug(publish_reports)
    video_by_slug = _by_slug(video_reports)
    rows: list[dict[str, Any]] = []
    for index, selected in enumerate(selected_rows, 1):
        slug = str(selected.get("slug", ""))
        publish = publish_by_slug.get(slug, {})
        video = video_by_slug.get(slug, {})
        rows.append(
            {
                "rank": selected.get("rank") or index,
                "topic": selected.get("topic", ""),
                "slug": slug,
                "final_score": selected.get("final_score", ""),
                "priority": selected.get("priority", ""),
                "action": selected.get("decision") or selected.get("action", ""),
                "article_type": selected.get("article_type", ""),
                "target_keyword": selected.get("target_keyword") or str(selected.get("topic", "")).lower(),
                "duplicate_risk": selected.get("risk_note") or selected.get("reason", ""),
                "estimated_value": selected.get("estimated_value", ""),
                "publish_status": publish.get("status", "not_processed"),
                "published_url": publish.get("published_url") or publish.get("article_url", ""),
                "video_package_path": video.get("video_package_path") or video.get("video_folder", ""),
                "next_action": _next_action(selected, publish, video),
                "note": publish.get("error") or video.get("error") or selected.get("reason", ""),
            }
        )
    return rows


def _top_rows(source_rows: list[dict[str, Any]], score_key: str, action_key: str = "decision") -> list[dict[str, Any]]:
    rows = sorted(source_rows, key=lambda row: numeric(row.get(score_key)), reverse=True)[:10]
    return [
        {
            "rank": index,
            "topic": row.get("topic", ""),
            "slug": row.get("slug", ""),
            "score": row.get(score_key, ""),
            "action": row.get(action_key, row.get("recommended_action", "")),
        }
        for index, row in enumerate(rows, 1)
    ]


def write_morning_outputs(
    selected_rows: list[dict[str, Any]],
    publish_reports: list[dict[str, Any]],
    video_reports: list[dict[str, Any]],
) -> dict[str, Any]:
    morning_rows = build_morning_command_rows(selected_rows, publish_reports, video_reports)
    publish_by_slug = _by_slug(publish_reports)
    video_by_slug = _by_slug(video_reports)
    for row in publish_reports:
        slug = str(row.get("slug", ""))
        row["video_package_path"] = video_by_slug.get(slug, {}).get("video_package_path", row.get("video_package_path", ""))
    write_csv(DATA_DIR / "morning_command_center.csv", morning_rows, MORNING_COMMAND_FIELDS)
    write_json(DATA_DIR / "morning_command_center.json", morning_rows)
    write_csv(DATA_DIR / "published_today.csv", publish_reports, PUBLISHED_TODAY_FIELDS)
    write_json(DATA_DIR / "published_today.json", publish_reports)

    selected_by_slug = _by_slug(selected_rows)
    article_rows = []
    for row in publish_reports:
        slug = str(row.get("slug", ""))
        article_rows.append(
            {
                "run_timestamp": row.get("run_timestamp", ""),
                "topic": row.get("topic", ""),
                "slug": slug,
                "article_type": row.get("article_type", selected_by_slug.get(slug, {}).get("article_type", "")),
                "status": row.get("status", ""),
                "markdown_path": row.get("markdown_path", ""),
                "json_path": row.get("json_path", ""),
                "word_count": row.get("word_count", ""),
                "youtube_embed_position": row.get("youtube_embed_position", ""),
                "error": row.get("error", ""),
            }
        )
    write_csv(DATA_DIR / "article_draft_report.csv", article_rows, ARTICLE_DRAFT_DAILY_FIELDS)
    write_json(DATA_DIR / "article_draft_report.json", article_rows)

    write_csv(DATA_DIR / "video_package_report.csv", video_reports, VIDEO_PACKAGE_DAILY_FIELDS)
    write_json(DATA_DIR / "video_package_report.json", video_reports)
    ensure_front_workbook_sheets(morning_rows, publish_reports, article_rows, video_reports)
    return {
        "morning_rows": len(morning_rows),
        "published_today": len(publish_reports),
        "article_draft_report": len(article_rows),
        "video_package_report": len(video_reports),
    }


def ensure_front_workbook_sheets(
    morning_rows: list[dict[str, Any]] | None = None,
    publish_rows: list[dict[str, Any]] | None = None,
    article_rows: list[dict[str, Any]] | None = None,
    video_rows: list[dict[str, Any]] | None = None,
    workbook_path: Path | None = None,
) -> bool:
    workbook_path = workbook_path or DATA_DIR / "master_dashboard.xlsx"
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return False

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    sheet_payloads = {
        "Morning Command Center": (morning_rows or read_csv(DATA_DIR / "morning_command_center.csv"), MORNING_COMMAND_FIELDS),
        "Published Today": (publish_rows or read_csv(DATA_DIR / "published_today.csv"), PUBLISHED_TODAY_FIELDS),
        "Article Draft Report": (article_rows or read_csv(DATA_DIR / "article_draft_report.csv"), ARTICLE_DRAFT_DAILY_FIELDS),
        "Video Package Report": (video_rows or read_csv(DATA_DIR / "video_package_report.csv"), VIDEO_PACKAGE_DAILY_FIELDS),
    }

    used_names: set[str] = set()
    for existing_name in list(workbook.sheetnames):
        safe_name = clean_sheet_name(existing_name, used_names)
        if safe_name != existing_name:
            workbook[existing_name].title = safe_name

    for sheet_name, fields in DEFAULT_SHEET_FIELDS.items():
        if sheet_name not in workbook.sheetnames:
            ws = workbook.create_sheet(sheet_name)
            ws.append([clean_excel_value(field) for field in fields])

    for sheet_name, (rows, fields) in sheet_payloads.items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        ws.append([clean_excel_value(field) for field in fields])
        for row in rows:
            ws.append([clean_excel_value(row.get(field, "")) for field in fields])

    priority_rows = read_csv(DATA_DIR / "today_selected_topics.csv")
    top_sheet_payloads = {
        "Top SEO": (_top_rows(priority_rows, "final_score"), ["rank", "topic", "slug", "score", "action"]),
        "Top Revenue": (_top_rows(priority_rows, "estimated_value"), ["rank", "topic", "slug", "score", "action"]),
        "Top Video": (_top_rows(priority_rows, "youtube_score"), ["rank", "topic", "slug", "score", "action"]),
        "Top Social": (_top_rows(priority_rows, "social_score"), ["rank", "topic", "slug", "score", "action"]),
    }
    for sheet_name, (rows, fields) in top_sheet_payloads.items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        ws.append([clean_excel_value(field) for field in fields])
        for row in rows:
            ws.append([clean_excel_value(row.get(field, "")) for field in fields])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 70)

    front = [workbook[name] for name in FRONT_SHEETS if name in workbook.sheetnames]
    remaining = [ws for ws in workbook.worksheets if ws.title not in FRONT_SHEETS]
    workbook._sheets = front + remaining
    result = safe_save_workbook(workbook, workbook_path)
    return bool(result.get("saved") or result.get("pending"))
