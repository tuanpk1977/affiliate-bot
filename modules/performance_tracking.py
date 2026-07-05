from __future__ import annotations

import csv
import html as html_lib
import json
import os
import re
import shutil
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
BASE_URL = "https://smileaireviewhub.com"

LIFECYCLE_FIELDS = [
    "slug",
    "topic",
    "article_url",
    "publish_date",
    "index_status",
    "google_clicks",
    "impressions",
    "ctr",
    "avg_position",
    "affiliate_clicks",
    "revenue_estimate",
    "youtube_url",
    "youtube_views",
    "social_status",
    "refresh_score",
    "next_action",
]

GSC_PAGE_FIELDS = ["slug", "page", "clicks", "impressions", "ctr", "position", "country", "query"]
GSC_QUERY_FIELDS = ["query", "page", "slug", "clicks", "impressions", "ctr", "position", "country"]
YOUTUBE_FIELDS = ["slug", "video_url", "views", "impressions", "ctr", "average_view_duration", "watch_time", "subscribers_gained"]
SOCIAL_FIELDS = ["date", "platform", "slug", "post_url", "views", "likes", "comments", "shares", "saves", "clicks", "notes"]
REVENUE_FIELDS = ["slug", "article_url", "topic", "affiliate_clicks", "revenue_estimate", "program", "google_clicks", "buyer_intent_score", "revenue_opportunity", "notes"]
REFRESH_FIELDS = ["slug", "reason", "priority", "suggested_action", "expected_impact"]
INTERNAL_LINK_FIELDS = ["source_slug", "source_url", "target_slug", "target_url", "anchor_text", "reason", "score", "status"]
COMPETITOR_TARGET_FIELDS = ["slug", "target_keyword", "competitor_url", "notes"]
COMPETITOR_GAP_FIELDS = ["slug", "target_keyword", "competitor_url", "missing_sections", "content_gap", "schema_gap", "video_gap", "affiliate_gap", "priority"]
SOCIAL_SUMMARY_FIELDS = ["platform", "posts", "views", "likes", "comments", "shares", "saves", "clicks", "click_potential"]


def slugify(value: str) -> str:
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", str(value or "").lower())).strip("-")


def slug_from_url(url: str) -> str:
    path = urlparse(str(url or "")).path.strip("/")
    if not path:
        return "home"
    parts = [part for part in path.split("/") if part not in {"review", "reviews", "compare", "category", "vi"}]
    return slugify(parts[-1] if parts else path)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def numeric(value: Any, default: float = 0.0) -> float:
    try:
        text = str(value).strip().replace("%", "")
        return float(text)
    except (TypeError, ValueError):
        return default


def load_json_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("topics", "scores", "todays_top_10", "selected_topics"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
    return []


def load_upload_links() -> dict[str, dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in (DATA_DIR / "upload_links.csv", ROOT / "video_output" / "upload_links.csv"):
        rows.extend(read_csv(path))
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        slug = slugify(row.get("FolderName") or row.get("folder") or row.get("slug") or slug_from_url(row.get("PageUrl", "")))
        if not slug:
            continue
        result[slug] = {
            "article_url": row.get("PageUrl", ""),
            "youtube_url": row.get("YoutubeVideoUrl", ""),
            "status": row.get("UploadStatus") or row.get("Status") or row.get("Notes") or "",
        }
    return result


def parse_gsc_export(input_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = read_csv(input_path)
    page_rows: list[dict[str, Any]] = []
    query_rows: list[dict[str, Any]] = []
    for row in rows:
        lowered = {str(k).strip().lower(): v for k, v in row.items()}
        page = lowered.get("page") or lowered.get("pages") or lowered.get("url") or lowered.get("landing page") or ""
        query = lowered.get("query") or lowered.get("queries") or ""
        country = lowered.get("country") or ""
        clicks = numeric(lowered.get("clicks"))
        impressions = numeric(lowered.get("impressions"))
        ctr = numeric(lowered.get("ctr"))
        position = numeric(lowered.get("position") or lowered.get("avg position"))
        slug = slug_from_url(page)
        record = {
            "slug": slug,
            "page": page,
            "clicks": clicks,
            "impressions": impressions,
            "ctr": ctr,
            "position": position,
            "country": country,
            "query": query,
        }
        page_rows.append(record)
        query_rows.append({"query": query, "page": page, "slug": slug, "clicks": clicks, "impressions": impressions, "ctr": ctr, "position": position, "country": country})
    return page_rows, query_rows


def parse_youtube_export(input_path: Path) -> list[dict[str, Any]]:
    rows = read_csv(input_path)
    parsed: list[dict[str, Any]] = []
    for row in rows:
        lowered = {str(k).strip().lower(): v for k, v in row.items()}
        slug = slugify(lowered.get("slug") or lowered.get("folder") or lowered.get("foldername") or "")
        video_url = lowered.get("video url") or lowered.get("url") or lowered.get("youtube url") or ""
        if not slug:
            slug = slug_from_url(video_url or lowered.get("title", ""))
        parsed.append(
            {
                "slug": slug,
                "video_url": video_url,
                "views": numeric(lowered.get("views")),
                "impressions": numeric(lowered.get("impressions")),
                "ctr": numeric(lowered.get("impressions click-through rate (%)") or lowered.get("ctr")),
                "average_view_duration": lowered.get("average view duration") or lowered.get("avg view duration") or "",
                "watch_time": numeric(lowered.get("watch time (hours)") or lowered.get("watch time")),
                "subscribers_gained": numeric(lowered.get("subscribers gained") or lowered.get("subscribers")),
            }
        )
    return parsed


def topic_map() -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in load_json_rows(DATA_DIR / "topic_scores.json"):
        slug = slugify(row.get("slug") or row.get("topic"))
        result[slug] = row
    for row in read_csv(DATA_DIR / "hottrend_latest_dashboard.csv"):
        slug = slugify(row.get("slug") or row.get("topic"))
        result.setdefault(slug, row)
    return result


def build_content_lifecycle() -> list[dict[str, Any]]:
    topics = topic_map()
    uploads = load_upload_links()
    gsc_rows = read_csv(DATA_DIR / "gsc_performance_pages.csv")
    youtube_rows = read_csv(DATA_DIR / "youtube_analytics.csv")
    social_rows = read_csv(DATA_DIR / "social_analytics.csv")
    slugs = set(topics) | set(uploads) | {row.get("slug", "") for row in gsc_rows} | {row.get("slug", "") for row in youtube_rows} | {row.get("slug", "") for row in social_rows}
    gsc_by_slug = aggregate_gsc(gsc_rows)
    youtube_by_slug = aggregate_youtube(youtube_rows)
    social_by_slug = aggregate_social(social_rows)
    rows: list[dict[str, Any]] = []
    for slug in sorted(filter(None, slugs)):
        topic = topics.get(slug, {}).get("topic") or slug.replace("-", " ").title()
        upload = uploads.get(slug, {})
        gsc = gsc_by_slug.get(slug, {})
        youtube = youtube_by_slug.get(slug, {})
        social = social_by_slug.get(slug, {})
        article_url = upload.get("article_url") or topics.get(slug, {}).get("article_url") or f"{BASE_URL}/{slug}/"
        clicks = numeric(gsc.get("clicks"))
        impressions = numeric(gsc.get("impressions"))
        affiliate_clicks = numeric(topics.get(slug, {}).get("affiliate_clicks"), default=round(clicks * 0.03, 1) if clicks else 0)
        revenue = numeric(topics.get(slug, {}).get("revenue_estimate"), default=round(affiliate_clicks * 2.0, 2))
        refresh_score = calculate_refresh_score(gsc, youtube, topics.get(slug, {}), upload)
        rows.append(
            {
                "slug": slug,
                "topic": topic,
                "article_url": article_url,
                "publish_date": topics.get(slug, {}).get("publish_date", ""),
                "index_status": "Needs GSC data" if not gsc else "Tracked",
                "google_clicks": clicks,
                "impressions": impressions,
                "ctr": gsc.get("ctr", ""),
                "avg_position": gsc.get("position", ""),
                "affiliate_clicks": affiliate_clicks,
                "revenue_estimate": revenue,
                "youtube_url": upload.get("youtube_url") or youtube.get("video_url", ""),
                "youtube_views": youtube.get("views", ""),
                "social_status": social.get("status", "No social data"),
                "refresh_score": refresh_score,
                "next_action": next_lifecycle_action(refresh_score, upload, gsc, topics.get(slug, {})),
            }
        )
    return rows


def aggregate_gsc(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {"clicks": 0.0, "impressions": 0.0, "weighted_position": 0.0, "ctr": 0.0})
    for row in rows:
        slug = row.get("slug", "")
        impressions = numeric(row.get("impressions"))
        grouped[slug]["clicks"] += numeric(row.get("clicks"))
        grouped[slug]["impressions"] += impressions
        grouped[slug]["weighted_position"] += numeric(row.get("position")) * max(impressions, 1)
    for slug, row in grouped.items():
        impressions = row["impressions"]
        row["ctr"] = round(row["clicks"] / impressions * 100, 2) if impressions else 0
        row["position"] = round(row["weighted_position"] / max(impressions, 1), 1)
    return dict(grouped)


def aggregate_youtube(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {"views": 0.0, "impressions": 0.0, "watch_time": 0.0, "subscribers_gained": 0.0, "video_url": ""})
    for row in rows:
        slug = row.get("slug", "")
        grouped[slug]["views"] += numeric(row.get("views"))
        grouped[slug]["impressions"] += numeric(row.get("impressions"))
        grouped[slug]["watch_time"] += numeric(row.get("watch_time"))
        grouped[slug]["subscribers_gained"] += numeric(row.get("subscribers_gained"))
        grouped[slug]["video_url"] = grouped[slug]["video_url"] or row.get("video_url", "")
    return dict(grouped)


def aggregate_social(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {"clicks": 0.0, "views": 0.0, "status": "Tracked"})
    for row in rows:
        slug = row.get("slug", "")
        grouped[slug]["clicks"] += numeric(row.get("clicks"))
        grouped[slug]["views"] += numeric(row.get("views"))
    return dict(grouped)


def calculate_refresh_score(gsc: dict[str, Any], youtube: dict[str, Any], topic: dict[str, Any], upload: dict[str, Any]) -> int:
    score = 0
    if numeric(gsc.get("impressions")) >= 100 and numeric(gsc.get("ctr")) < 1.5:
        score += 25
    if numeric(gsc.get("position")) > 12:
        score += 20
    if numeric(topic.get("buyer_intent_score") or topic.get("buyer_intent")) >= 65 and numeric(topic.get("revenue_estimate")) <= 0:
        score += 20
    if not upload.get("youtube_url"):
        score += 15
    if numeric(youtube.get("views")) < 10 and upload.get("youtube_url"):
        score += 10
    return min(score, 100)


def next_lifecycle_action(refresh_score: int, upload: dict[str, Any], gsc: dict[str, Any], topic: dict[str, Any]) -> str:
    if refresh_score >= 60:
        return "Refresh article"
    if not upload.get("youtube_url"):
        return "Create or upload video"
    if numeric(gsc.get("impressions")) > 100 and numeric(gsc.get("clicks")) < 3:
        return "Improve title/meta"
    if numeric(topic.get("revenue_score")) >= 70:
        return "Improve affiliate CTA"
    return "Monitor"


def build_revenue_dashboard(lifecycle: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in lifecycle:
        clicks = numeric(row.get("google_clicks"))
        revenue = numeric(row.get("revenue_estimate"))
        affiliate_clicks = numeric(row.get("affiliate_clicks"))
        opportunity = "High traffic low revenue" if clicks >= 50 and revenue < 5 else "Low traffic high intent" if clicks < 10 and affiliate_clicks > 0 else "Monitor"
        rows.append(
            {
                "slug": row.get("slug", ""),
                "article_url": row.get("article_url", ""),
                "topic": row.get("topic", ""),
                "affiliate_clicks": affiliate_clicks,
                "revenue_estimate": revenue,
                "program": infer_program(row.get("slug", "")),
                "google_clicks": clicks,
                "buyer_intent_score": "",
                "revenue_opportunity": opportunity,
                "notes": "Estimated from available CSV/log data when real affiliate data is missing.",
            }
        )
    return sorted(rows, key=lambda item: (numeric(item.get("revenue_estimate")), numeric(item.get("google_clicks"))), reverse=True)


def build_social_dashboard(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"platform": "", "posts": 0, "views": 0.0, "likes": 0.0, "comments": 0.0, "shares": 0.0, "saves": 0.0, "clicks": 0.0}
    )
    for row in rows:
        platform = str(row.get("platform") or "unknown").strip() or "unknown"
        grouped[platform]["platform"] = platform
        grouped[platform]["posts"] += 1
        for field in ("views", "likes", "comments", "shares", "saves", "clicks"):
            grouped[platform][field] += numeric(row.get(field))
    result = []
    for row in grouped.values():
        click_potential = row["clicks"] or round((row["likes"] + row["comments"] * 2 + row["shares"] * 3 + row["saves"] * 2) * 0.05, 1)
        result.append({**row, "click_potential": click_potential})
    return sorted(result, key=lambda item: (numeric(item.get("click_potential")), numeric(item.get("views"))), reverse=True)


def build_competitor_gap_rows(targets: list[dict[str, Any]], lifecycle: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lifecycle_by_slug = {row.get("slug", ""): row for row in lifecycle}
    rows: list[dict[str, Any]] = []
    for target in targets:
        slug = slugify(target.get("slug") or target.get("target_keyword") or slug_from_url(target.get("competitor_url", "")))
        current = lifecycle_by_slug.get(slug, {})
        missing = []
        if not current.get("youtube_url"):
            missing.append("video")
        if numeric(current.get("google_clicks")) <= 0:
            missing.append("GSC traction")
        priority = "High" if missing and target.get("competitor_url") else "Medium" if missing else "Low"
        rows.append(
            {
                "slug": slug,
                "target_keyword": target.get("target_keyword", ""),
                "competitor_url": target.get("competitor_url", ""),
                "missing_sections": "; ".join(missing),
                "content_gap": "Manual review needed against competitor page." if target.get("competitor_url") else "Add competitor URL to evaluate.",
                "schema_gap": "Check Article/FAQ/Review schema manually.",
                "video_gap": "Missing or unlinked video." if "video" in missing else "Video linked or not assessed.",
                "affiliate_gap": "Review CTA and affiliate placement manually.",
                "priority": priority,
            }
        )
    return rows


def infer_program(slug: str) -> str:
    parts = slug.split("-")
    return parts[0].title() if parts else "Unknown"


def recommend_refresh(lifecycle: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in lifecycle:
        score = numeric(row.get("refresh_score"))
        reasons = []
        if numeric(row.get("impressions")) >= 100 and numeric(row.get("ctr")) < 1.5:
            reasons.append("High impressions but low CTR")
        if numeric(row.get("avg_position")) > 12:
            reasons.append("Average position needs improvement")
        if not row.get("youtube_url"):
            reasons.append("Missing YouTube video")
        if numeric(row.get("revenue_estimate")) <= 0 and numeric(row.get("google_clicks")) > 0:
            reasons.append("Traffic exists but revenue is low")
        if score or reasons:
            priority = "High" if score >= 60 else "Medium" if score >= 30 else "Low"
            rows.append(
                {
                    "slug": row.get("slug", ""),
                    "reason": "; ".join(reasons) or "Periodic content refresh candidate",
                    "priority": priority,
                    "suggested_action": row.get("next_action", "Refresh article"),
                    "expected_impact": "Improve CTR, rankings, affiliate clicks, or video coverage.",
                }
            )
    return sorted(rows, key=lambda item: {"High": 3, "Medium": 2, "Low": 1}.get(item["priority"], 0), reverse=True)


def recommend_internal_links(lifecycle: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for source in lifecycle:
        source_tokens = set(source.get("slug", "").split("-"))
        for target in lifecycle:
            if source["slug"] == target["slug"]:
                continue
            target_tokens = set(target.get("slug", "").split("-"))
            overlap = len(source_tokens & target_tokens)
            if overlap < 2:
                continue
            rows.append(
                {
                    "source_slug": source.get("slug", ""),
                    "source_url": source.get("article_url", ""),
                    "target_slug": target.get("slug", ""),
                    "target_url": target.get("article_url", ""),
                    "anchor_text": target.get("topic", ""),
                    "reason": f"Shared topic tokens: {', '.join(sorted(source_tokens & target_tokens))}",
                    "score": overlap * 10,
                    "status": "dry_run",
                }
            )
    return sorted(rows, key=lambda row: numeric(row.get("score")), reverse=True)


def ensure_template(path: Path, fields: list[str]) -> None:
    if not path.exists():
        write_csv(path, [], fields)


INVALID_SHEET_CHARS = re.compile(r"[\\/\?\*\[\]:]")
INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def clean_sheet_name(name: str, used: set[str] | None = None) -> str:
    used = used or set()
    cleaned = INVALID_SHEET_CHARS.sub(" ", str(name or "Sheet")).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def clean_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = INVALID_XML_CHARS.sub("", value)
        if cleaned.startswith("="):
            return "'" + cleaned
        return cleaned
    return value


def workbook_is_locked(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        with path.open("r+b"):
            return False
    except OSError:
        return True


def validate_workbook_file(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "path": str(path),
        "exists": path.exists(),
        "valid": False,
        "errors": [],
        "sheet_count": 0,
        "worksheet_names": [],
        "row_counts": {},
        "file_size": path.stat().st_size if path.exists() else 0,
    }
    if not path.exists():
        result["errors"].append("Workbook file does not exist.")
        return result
    try:
        from openpyxl import load_workbook
    except ImportError:
        result["errors"].append("openpyxl is not installed.")
        return result
    try:
        workbook = load_workbook(path)
    except Exception as exc:
        result["errors"].append(f"openpyxl could not load workbook: {exc}")
        return result
    names = workbook.sheetnames
    result["sheet_count"] = len(names)
    result["worksheet_names"] = names
    result["row_counts"] = {ws.title: ws.max_row for ws in workbook.worksheets}
    if len(names) != len(set(names)):
        result["errors"].append("Duplicate worksheet names detected.")
    for name in names:
        if len(name) > 31:
            result["errors"].append(f"Worksheet name too long: {name}")
        if INVALID_SHEET_CHARS.search(name):
            result["errors"].append(f"Worksheet name contains illegal Excel characters: {name}")
    for ws in workbook.worksheets:
        for merged_range in ws.merged_cells.ranges:
            if not str(merged_range):
                result["errors"].append(f"Invalid merged range in {ws.title}.")
        for table_name, table in ws.tables.items():
            table_ref = getattr(table, "ref", table if isinstance(table, str) else "")
            if not table_ref:
                result["errors"].append(f"Invalid table range in {ws.title}: {table_name}")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and INVALID_XML_CHARS.search(value):
                    result["errors"].append(f"Invalid XML character in {ws.title}!{cell.coordinate}")
                if isinstance(value, str) and value.startswith("=") and len(value) == 1:
                    result["errors"].append(f"Broken formula in {ws.title}!{cell.coordinate}")
                if cell.hyperlink and cell.hyperlink.target and INVALID_XML_CHARS.search(str(cell.hyperlink.target)):
                    result["errors"].append(f"Invalid hyperlink in {ws.title}!{cell.coordinate}")
    result["valid"] = not result["errors"]
    return result


def strip_excel_objects_for_safe_mode(workbook: Any) -> None:
    for ws in workbook.worksheets:
        if hasattr(ws, "_tables"):
            try:
                ws._tables.clear()
            except Exception:
                ws._tables = {}
        if hasattr(ws, "_charts"):
            ws._charts = []


def safe_save_workbook(workbook: Any, workbook_path: Path) -> dict[str, Any]:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    strip_excel_objects_for_safe_mode(workbook)
    backup_dir = workbook_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = timestamp()
    tmp_path = workbook_path.with_name(f"{workbook_path.stem}.tmp.xlsx")
    pending_path = workbook_path.with_name(f"{workbook_path.stem}.pending_{stamp}.xlsx")
    backup_path = backup_dir / f"{workbook_path.stem}_{stamp}.xlsx"

    target_path = pending_path if workbook_is_locked(workbook_path) else workbook_path
    if target_path == pending_path:
        workbook.save(pending_path)
        validation = validate_workbook_file(pending_path)
        print(f"WARNING: {workbook_path} appears to be open in Excel. Wrote pending workbook instead: {pending_path}")
        return {
            "saved": False,
            "pending": True,
            "path": str(pending_path),
            "backup_path": "",
            "validation": validation,
            "excel_safe_mode": True,
        }

    if tmp_path.exists():
        try:
            tmp_path.unlink()
        except OSError:
            pass
    workbook.save(tmp_path)
    tmp_validation = validate_workbook_file(tmp_path)
    if not tmp_validation["valid"]:
        return {
            "saved": False,
            "pending": False,
            "path": str(tmp_path),
            "backup_path": "",
            "validation": tmp_validation,
            "excel_safe_mode": True,
        }
    if workbook_path.exists():
        shutil.copy2(workbook_path, backup_path)
    os.replace(tmp_path, workbook_path)
    final_validation = validate_workbook_file(workbook_path)
    return {
        "saved": final_validation["valid"],
        "pending": False,
        "path": str(workbook_path),
        "backup_path": str(backup_path) if backup_path.exists() else "",
        "validation": final_validation,
        "excel_safe_mode": True,
    }


def print_workbook_report(save_result: dict[str, Any]) -> None:
    validation = save_result.get("validation", {})
    print(f"Workbook path: {save_result.get('path', validation.get('path', ''))}")
    print(f"File size: {validation.get('file_size', 0)} bytes")
    print(f"Worksheet count: {validation.get('sheet_count', 0)}")
    names = validation.get("worksheet_names", [])
    print("Worksheet names: " + (", ".join(names) if names else "(none)"))
    row_counts = validation.get("row_counts", {})
    if row_counts:
        print("Row counts per sheet:")
        for name, count in row_counts.items():
            print(f"  - {name}: {count}")
    print("Validation result: " + ("PASS" if validation.get("valid") else "FAIL"))
    if validation.get("errors"):
        for error in validation["errors"]:
            print(f"  ERROR: {error}")
    print(f"Backup path: {save_result.get('backup_path') or '(none)'}")
    print(f"Excel-safe mode: {'YES' if save_result.get('excel_safe_mode') else 'NO'}")


def update_master_workbook(sheets: dict[str, tuple[list[dict[str, Any]], list[str]]], path: Path | None = None) -> bool:
    workbook_path = path or DATA_DIR / "master_dashboard.xlsx"
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return False
    if workbook_path.exists():
        try:
            workbook = load_workbook(workbook_path)
        except Exception:
            backup = workbook_path.with_suffix(f".corrupt-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx")
            shutil.copy2(workbook_path, backup)
            workbook = Workbook()
            workbook.remove(workbook.active)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
    existing_used_sheet_names = set()
    for existing_name in list(workbook.sheetnames):
        safe_existing = clean_sheet_name(existing_name, existing_used_sheet_names)
        if safe_existing != existing_name:
            workbook[existing_name].title = safe_existing
    new_sheet_names: set[str] = set()
    for sheet_name, (rows, fields) in sheets.items():
        safe_name = clean_sheet_name(sheet_name, new_sheet_names)
        if safe_name in workbook.sheetnames:
            del workbook[safe_name]
        ws = workbook.create_sheet(safe_name)
        ws.append([clean_excel_value(field) for field in fields])
        for row in rows:
            ws.append([clean_excel_value(row.get(field, "")) for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 60)
    result = safe_save_workbook(workbook, workbook_path)
    print_workbook_report(result)
    return bool(result.get("saved") or result.get("pending"))


def write_business_outputs() -> dict[str, Any]:
    lifecycle = build_content_lifecycle()
    revenue = build_revenue_dashboard(lifecycle)
    refresh = recommend_refresh(lifecycle)
    internal = recommend_internal_links(lifecycle)
    social_rows = read_csv(DATA_DIR / "social_analytics.csv")
    social_dashboard = build_social_dashboard(social_rows)
    youtube_rows = read_csv(DATA_DIR / "youtube_analytics.csv")
    competitor_targets = read_csv(DATA_DIR / "competitor_targets.csv")
    competitor_gaps = build_competitor_gap_rows(competitor_targets, lifecycle)
    write_csv(DATA_DIR / "content_lifecycle.csv", lifecycle, LIFECYCLE_FIELDS)
    write_json(DATA_DIR / "content_lifecycle.json", lifecycle)
    write_csv(DATA_DIR / "revenue_dashboard.csv", revenue, REVENUE_FIELDS)
    write_json(DATA_DIR / "revenue_dashboard.json", revenue)
    write_csv(DATA_DIR / "content_refresh_recommendations.csv", refresh, REFRESH_FIELDS)
    write_json(DATA_DIR / "content_refresh_recommendations.json", refresh)
    write_csv(DATA_DIR / "internal_link_recommendations.csv", internal, INTERNAL_LINK_FIELDS)
    ensure_template(DATA_DIR / "social_analytics.csv", SOCIAL_FIELDS)
    write_csv(DATA_DIR / "social_analytics_dashboard.csv", social_dashboard, SOCIAL_SUMMARY_FIELDS)
    ensure_template(DATA_DIR / "competitor_targets.csv", COMPETITOR_TARGET_FIELDS)
    write_csv(DATA_DIR / "competitor_gap_analysis.csv", competitor_gaps, COMPETITOR_GAP_FIELDS)
    excel = update_master_workbook(
        {
            "Content Lifecycle": (lifecycle, LIFECYCLE_FIELDS),
            "Revenue Dashboard": (revenue, REVENUE_FIELDS),
            "YouTube Analytics": (youtube_rows, YOUTUBE_FIELDS),
            "Social Analytics": (social_rows, SOCIAL_FIELDS),
            "Social Summary": (social_dashboard, SOCIAL_SUMMARY_FIELDS),
            "Refresh Recommendations": (refresh, REFRESH_FIELDS),
            "Competitor Gaps": (competitor_gaps, COMPETITOR_GAP_FIELDS),
            "Internal Link Ideas": (internal, INTERNAL_LINK_FIELDS),
        }
    )
    executive = build_daily_executive_dashboard(lifecycle, revenue, refresh)
    write_json(DATA_DIR / "daily_executive_dashboard.json", executive)
    write_html_executive_dashboard(DATA_DIR / "daily_executive_dashboard.html", executive)
    return {
        "lifecycle": len(lifecycle),
        "revenue": len(revenue),
        "refresh": len(refresh),
        "internal_links": len(internal),
        "social": len(social_dashboard),
        "competitor_gaps": len(competitor_gaps),
        "excel": excel,
    }


def build_daily_executive_dashboard(lifecycle: list[dict[str, Any]], revenue: list[dict[str, Any]], refresh: list[dict[str, Any]]) -> dict[str, Any]:
    topic_scores = load_json_rows(DATA_DIR / "topic_scores.json")
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "todays_top_10_topics": topic_scores[:10],
        "top_3_topics_to_write": [row for row in topic_scores if numeric(row.get("total_score")) >= 70][:3],
        "top_3_video_candidates": [row for row in topic_scores if "YouTube" in str(row.get("content_decision", ""))][:3],
        "top_3_refresh_candidates": refresh[:3],
        "revenue_opportunities": revenue[:10],
        "gsc_opportunities": [row for row in lifecycle if numeric(row.get("impressions")) >= 100 and numeric(row.get("ctr")) < 1.5][:10],
        "youtube_opportunities": [row for row in lifecycle if not row.get("youtube_url")][:10],
        "social_draft_opportunities": [row for row in topic_scores if numeric(row.get("total_score")) >= 80][:10],
        "warnings": [row for row in topic_scores if numeric(row.get("total_score")) < 60][:10],
    }


def write_html_executive_dashboard(path: Path, dashboard: dict[str, Any]) -> None:
    def table(rows: list[dict[str, Any]], fields: list[str]) -> str:
        header = "".join(f"<th>{html_lib.escape(field)}</th>" for field in fields)
        body = []
        for row in rows:
            body.append("<tr>" + "".join(f"<td>{html_lib.escape(str(row.get(field, '')))}</td>" for field in fields) + "</tr>")
        return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body)}</tbody></table>"

    html = f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>Daily Executive Dashboard</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;background:#f8fafc;color:#172033}}table{{border-collapse:collapse;width:100%;background:#fff;margin:12px 0 28px}}td,th{{border:1px solid #d8e1ea;padding:8px;text-align:left}}th{{background:#e6fffb}}h1,h2{{color:#0f766e}}</style></head>
<body>
<h1>Daily Executive Dashboard</h1>
<p>Generated: {html_lib.escape(str(dashboard.get('generated_at', '')))}</p>
<h2>Today's Top 10 Topics</h2>{table(dashboard.get('todays_top_10_topics', []), ['topic','total_score','recommendation','content_decision'])}
<h2>Top 3 Topics To Write</h2>{table(dashboard.get('top_3_topics_to_write', []), ['topic','total_score','recommendation','content_decision'])}
<h2>Top 3 Video Candidates</h2>{table(dashboard.get('top_3_video_candidates', []), ['topic','total_score','video_priority','content_decision'])}
<h2>Top 3 Refresh Candidates</h2>{table(dashboard.get('top_3_refresh_candidates', []), REFRESH_FIELDS)}
<h2>Revenue Opportunities</h2>{table(dashboard.get('revenue_opportunities', []), REVENUE_FIELDS[:8])}
<h2>Warnings / Risky Topics</h2>{table(dashboard.get('warnings', []), ['topic','total_score','recommendation','reason'])}
</body></html>"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
