from __future__ import annotations

import argparse
import io
import json
import logging
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from google.auth.transport.requests import Request
    from google.oauth2 import service_account
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload
except ImportError:
    Request = None
    service_account = None
    Credentials = None
    InstalledAppFlow = None
    build = None
    HttpError = Exception
    MediaIoBaseDownload = None


SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class SheetSummary:
    name: str
    status: str = "checked"
    changed_cells: list[str] = field(default_factory=list)
    formula_changes: list[str] = field(default_factory=list)
    hyperlink_changes: list[str] = field(default_factory=list)
    new_rows: int = 0
    deleted_rows: int = 0
    image_change: str = "not detected"


@dataclass
class CompareResult:
    old_exists: bool
    sheets_checked: int = 0
    new_sheets: list[str] = field(default_factory=list)
    deleted_sheets: list[str] = field(default_factory=list)
    sheet_summaries: list[SheetSummary] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def changed_cell_count(self) -> int:
        return sum(len(item.changed_cells) for item in self.sheet_summaries)

    @property
    def formula_change_count(self) -> int:
        return sum(len(item.formula_changes) for item in self.sheet_summaries)

    @property
    def hyperlink_change_count(self) -> int:
        return sum(len(item.hyperlink_changes) for item in self.sheet_summaries)

    @property
    def new_row_count(self) -> int:
        return sum(item.new_rows for item in self.sheet_summaries)

    @property
    def deleted_row_count(self) -> int:
        return sum(item.deleted_rows for item in self.sheet_summaries)


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M")


def now_human() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_config(base_dir: Path) -> dict[str, Any]:
    config_path = base_dir / "config.json"
    if not config_path.exists():
        raise FileNotFoundError(f"Missing config file: {config_path}")
    with config_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def ensure_folders(base_dir: Path, config: dict[str, Any]) -> dict[str, Path]:
    folders = {
        "reports": base_dir / config.get("reports_folder", "reports"),
        "snapshots": base_dir / config.get("snapshots_folder", "snapshots"),
        "logs": base_dir / config.get("logs_folder", "logs"),
    }
    for folder in folders.values():
        folder.mkdir(parents=True, exist_ok=True)
    return folders


def setup_logging(logs_dir: Path) -> Path:
    log_path = logs_dir / f"sheet_sync_{now_stamp()}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return log_path


def extract_spreadsheet_id(url_or_id: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url_or_id)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]+", url_or_id):
        return url_or_id
    raise ValueError("Could not extract spreadsheet ID from source_sheet_url.")


def public_export_url(file_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"


def looks_like_xlsx(content: bytes) -> bool:
    return content.startswith(b"PK")


def download_public_export(file_id: str, timeout: int = 60) -> bytes:
    url = public_export_url(file_id)
    logging.info("Trying public/link-access XLSX export.")
    response = requests.get(url, timeout=timeout, allow_redirects=True)
    response.raise_for_status()
    if not looks_like_xlsx(response.content):
        raise RuntimeError("Public export did not return an XLSX file.")
    return response.content


def google_libs_available() -> bool:
    return all([Request, service_account, Credentials, InstalledAppFlow, build, MediaIoBaseDownload])


def get_oauth_credentials(base_dir: Path, config: dict[str, Any]):
    credentials_path = base_dir / config.get("oauth_credentials_file", "credentials.json")
    token_path = base_dir / config.get("oauth_token_file", "token.json")
    if not credentials_path.exists():
        raise FileNotFoundError(f"OAuth credentials file not found: {credentials_path}")

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")
    return creds


def get_service_account_credentials(base_dir: Path, config: dict[str, Any]):
    service_account_path = base_dir / config.get("service_account_file", "service_account.json")
    if not service_account_path.exists():
        raise FileNotFoundError(f"Service account file not found: {service_account_path}")
    return service_account.Credentials.from_service_account_file(str(service_account_path), scopes=SCOPES)


def download_with_drive_api(file_id: str, base_dir: Path, config: dict[str, Any]) -> bytes:
    if not google_libs_available():
        raise RuntimeError("Google API libraries are not installed. Run: python -m pip install -r requirements.txt")

    auth_mode = config.get("auth_mode", "auto").lower()
    credential_errors: list[str] = []

    credential_attempts = []
    if auth_mode in {"auto", "oauth"}:
        credential_attempts.append(("OAuth", get_oauth_credentials))
    if auth_mode in {"auto", "service_account"}:
        credential_attempts.append(("service account", get_service_account_credentials))

    for label, factory in credential_attempts:
        try:
            logging.info("Trying Google Drive API export with %s credentials.", label)
            creds = factory(base_dir, config)
            service = build("drive", "v3", credentials=creds, cache_discovery=False)
            request = service.files().export_media(fileId=file_id, mimeType=XLSX_MIME)
            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            content = buffer.getvalue()
            if not looks_like_xlsx(content):
                raise RuntimeError("Drive API export did not return an XLSX file.")
            return content
        except Exception as exc:
            credential_errors.append(f"{label}: {exc}")
            logging.warning("Drive API export failed with %s credentials: %s", label, exc)

    raise RuntimeError("All authenticated export attempts failed. " + " | ".join(credential_errors))


def download_latest_xlsx(file_id: str, base_dir: Path, config: dict[str, Any]) -> tuple[bytes, list[str]]:
    notes: list[str] = []
    auth_mode = config.get("auth_mode", "auto").lower()
    if auth_mode in {"auto", "public"}:
        try:
            return download_public_export(file_id), ["Downloaded by public/link-access XLSX export."]
        except Exception as exc:
            message = f"Public/link-access export failed: {exc}"
            notes.append(message)
            logging.warning(message)
            if auth_mode == "public":
                raise

    try:
        content = download_with_drive_api(file_id, base_dir, config)
    except Exception as exc:
        if notes:
            raise RuntimeError("; ".join(notes) + " | " + str(exc)) from exc
        raise
    notes.append("Downloaded by Google Drive API read-only export.")
    return content, notes


def safe_load_workbook(path: Path):
    return load_workbook(path, data_only=False)


def cell_display(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    return str(value)


def hyperlink_target(cell) -> str:
    if cell.hyperlink is None:
        return ""
    return str(cell.hyperlink.target or cell.hyperlink.location or "")


def image_count(sheet: Worksheet) -> int | None:
    images = getattr(sheet, "_images", None)
    if images is None:
        return None
    return len(images)


def compare_sheet(old_sheet: Worksheet, new_sheet: Worksheet, max_examples: int = 200) -> SheetSummary:
    summary = SheetSummary(name=new_sheet.title)
    max_row = max(old_sheet.max_row, new_sheet.max_row)
    max_col = max(old_sheet.max_column, new_sheet.max_column)

    summary.new_rows = max(0, new_sheet.max_row - old_sheet.max_row)
    summary.deleted_rows = max(0, old_sheet.max_row - new_sheet.max_row)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            old_cell = old_sheet.cell(row=row, column=col)
            new_cell = new_sheet.cell(row=row, column=col)
            coord = new_cell.coordinate

            old_value = old_cell.value
            new_value = new_cell.value
            if old_value != new_value and len(summary.changed_cells) < max_examples:
                summary.changed_cells.append(f"{coord}: {old_value!r} -> {new_value!r}")

            old_formula = old_value if isinstance(old_value, str) and old_value.startswith("=") else ""
            new_formula = new_value if isinstance(new_value, str) and new_value.startswith("=") else ""
            if old_formula != new_formula and len(summary.formula_changes) < max_examples:
                summary.formula_changes.append(f"{coord}: {old_formula!r} -> {new_formula!r}")

            old_link = hyperlink_target(old_cell)
            new_link = hyperlink_target(new_cell)
            if old_link != new_link and len(summary.hyperlink_changes) < max_examples:
                summary.hyperlink_changes.append(f"{coord}: {old_link!r} -> {new_link!r}")

    old_images = image_count(old_sheet)
    new_images = image_count(new_sheet)
    if old_images is None or new_images is None:
        summary.image_change = "not accessible through openpyxl"
    elif old_images == new_images:
        summary.image_change = f"no count change ({new_images})"
    else:
        summary.image_change = f"{old_images} -> {new_images}"

    return summary


def compare_workbooks(old_path: Path | None, new_path: Path) -> CompareResult:
    if old_path is None or not old_path.exists():
        wb = safe_load_workbook(new_path)
        result = CompareResult(old_exists=False, sheets_checked=len(wb.sheetnames))
        result.new_sheets = list(wb.sheetnames)
        result.notes.append("Initial run: no previous local workbook existed.")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.sheet_summaries.append(
                SheetSummary(
                    name=sheet_name,
                    status="initial snapshot",
                    new_rows=ws.max_row,
                    image_change=f"{image_count(ws)} images detected" if image_count(ws) is not None else "not accessible",
                )
            )
        wb.close()
        return result

    old_wb = safe_load_workbook(old_path)
    new_wb = safe_load_workbook(new_path)
    result = CompareResult(old_exists=True)

    old_names = set(old_wb.sheetnames)
    new_names = set(new_wb.sheetnames)
    result.new_sheets = sorted(new_names - old_names)
    result.deleted_sheets = sorted(old_names - new_names)

    common = [name for name in new_wb.sheetnames if name in old_names]
    result.sheets_checked = len(common)
    for sheet_name in common:
        result.sheet_summaries.append(compare_sheet(old_wb[sheet_name], new_wb[sheet_name]))

    if result.new_sheets:
        for sheet_name in result.new_sheets:
            ws = new_wb[sheet_name]
            result.sheet_summaries.append(
                SheetSummary(name=sheet_name, status="new sheet", new_rows=ws.max_row)
            )

    old_wb.close()
    new_wb.close()
    result.notes.append("Formatting is preserved by replacing the local file with the latest Google XLSX export.")
    result.notes.append("Comments, notes, drawings, charts, and images are only reported when present in the exported XLSX and readable by openpyxl.")
    return result


def copy_snapshot(source: Path, snapshots_dir: Path, prefix: str) -> Path | None:
    if not source.exists():
        return None
    snapshot_path = snapshots_dir / f"{prefix}_{now_stamp()}.xlsx"
    shutil.copy2(source, snapshot_path)
    return snapshot_path


def prune_snapshots(snapshots_dir: Path, keep: int) -> None:
    if keep <= 0:
        return
    snapshots = sorted(snapshots_dir.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    for old_snapshot in snapshots[keep:]:
        try:
            old_snapshot.unlink()
        except OSError as exc:
            logging.warning("Could not prune snapshot %s: %s", old_snapshot, exc)


def write_report(
    report_path: Path,
    source_url: str,
    status: str,
    result: CompareResult | None,
    notes: list[str],
    errors: list[str],
    snapshot_path: Path | None,
    output_file: Path,
) -> None:
    lines = [
        "# Google Sheet Sync Report",
        "",
        f"- Run time: {now_human()}",
        f"- Source sheet URL: {source_url}",
        f"- Status: {status}",
        f"- Local output file: {output_file}",
        f"- Snapshot: {snapshot_path if snapshot_path else 'none'}",
        "",
    ]

    if result:
        lines.extend(
            [
                "## Totals",
                "",
                f"- Number of sheets checked: {result.sheets_checked}",
                f"- New sheets: {len(result.new_sheets)}",
                f"- Deleted sheets: {len(result.deleted_sheets)}",
                f"- New rows detected: {result.new_row_count}",
                f"- Deleted rows detected: {result.deleted_row_count}",
                f"- Changed cells: {result.changed_cell_count}",
                f"- Changed formulas: {result.formula_change_count}",
                f"- Changed links: {result.hyperlink_change_count}",
                "",
            ]
        )

        if result.new_sheets:
            lines.extend(["## New Sheets", "", *[f"- {name}" for name in result.new_sheets], ""])
        if result.deleted_sheets:
            lines.extend(["## Deleted Sheets", "", *[f"- {name}" for name in result.deleted_sheets], ""])

        lines.extend(["## Summary by Sheet", ""])
        for sheet in result.sheet_summaries:
            lines.extend(
                [
                    f"### {sheet.name}",
                    "",
                    f"- Status: {sheet.status}",
                    f"- New rows: {sheet.new_rows}",
                    f"- Deleted rows: {sheet.deleted_rows}",
                    f"- Changed cells shown: {len(sheet.changed_cells)}",
                    f"- Formula changes shown: {len(sheet.formula_changes)}",
                    f"- Link changes shown: {len(sheet.hyperlink_changes)}",
                    f"- Image count change: {sheet.image_change}",
                    "",
                ]
            )
            if sheet.changed_cells:
                lines.extend(["Changed cells:", ""])
                lines.extend([f"- {item}" for item in sheet.changed_cells[:50]])
                if len(sheet.changed_cells) > 50:
                    lines.append(f"- ... {len(sheet.changed_cells) - 50} more shown in internal comparison limit")
                lines.append("")
            if sheet.formula_changes:
                lines.extend(["Formula changes:", ""])
                lines.extend([f"- {item}" for item in sheet.formula_changes[:50]])
                lines.append("")
            if sheet.hyperlink_changes:
                lines.extend(["Link changes:", ""])
                lines.extend([f"- {item}" for item in sheet.hyperlink_changes[:50]])
                lines.append("")
    else:
        lines.extend(["## Totals", "", "- Number of sheets checked: 0", ""])

    all_notes = list(notes)
    if result:
        all_notes.extend(result.notes)
    if all_notes:
        lines.extend(["## Important Notes", ""])
        lines.extend([f"- {note}" for note in all_notes])
        lines.append("")

    if errors:
        lines.extend(["## Errors or Permissions Issues", ""])
        lines.extend([f"- {error}" for error in errors])
        lines.append("")
    else:
        lines.extend(["## Errors or Permissions Issues", "", "- None", ""])

    report_path.write_text("\n".join(lines), encoding="utf-8")


def atomic_replace(source: Path, destination: Path) -> None:
    try:
        if destination.exists():
            destination.unlink()
        source.replace(destination)
    except PermissionError as exc:
        raise PermissionError(f"Could not update {destination}. Close it in Excel and run again.") from exc


def run() -> int:
    parser = argparse.ArgumentParser(description="Read-only local sync for a shared Google Sheet.")
    parser.add_argument("--config", default="config.json", help="Config file name or path.")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = base_dir / config_path

    config = json.loads(config_path.read_text(encoding="utf-8"))
    folders = ensure_folders(base_dir, config)
    log_path = setup_logging(folders["logs"])
    logging.info("Log file: %s", log_path)

    source_url = config["source_sheet_url"]
    output_file = base_dir / config.get("output_file", "local_copy.xlsx")
    temp_file = base_dir / f".download_{now_stamp()}.xlsx"
    report_path: Path | None = None
    snapshot_path: Path | None = None
    result: CompareResult | None = None
    notes: list[str] = []
    errors: list[str] = []

    try:
        file_id = extract_spreadsheet_id(source_url)
        content, download_notes = download_latest_xlsx(file_id, base_dir, config)
        notes.extend(download_notes)
        temp_file.write_bytes(content)

        if output_file.exists():
            snapshot_path = copy_snapshot(output_file, folders["snapshots"], "backup_before_update")
            report_path = folders["reports"] / f"report_{now_stamp()}.md"
            result = compare_workbooks(output_file, temp_file)
        else:
            report_path = folders["reports"] / "initial_report.md"
            result = compare_workbooks(None, temp_file)

        atomic_replace(temp_file, output_file)
        if not snapshot_path:
            snapshot_path = copy_snapshot(output_file, folders["snapshots"], "initial_snapshot")

        prune_snapshots(folders["snapshots"], int(config.get("keep_snapshots", 30)))
        write_report(report_path, source_url, "success", result, notes, errors, snapshot_path, output_file)
        logging.info("Sync completed successfully. Report: %s", report_path)
        return 0

    except requests.exceptions.ConnectionError as exc:
        errors.append(f"No internet or connection failure: {exc}")
    except requests.exceptions.HTTPError as exc:
        errors.append(f"HTTP/export failure: {exc}")
    except HttpError as exc:
        errors.append(f"Google API failure or permission issue: {exc}")
    except PermissionError as exc:
        errors.append(str(exc))
    except Exception as exc:
        errors.append(f"Sync failed: {exc}")
        logging.exception("Sync failed.")
    finally:
        if temp_file.exists():
            try:
                temp_file.unlink()
            except OSError:
                pass

    if report_path is None:
        report_path = folders["reports"] / f"report_{now_stamp()}.md"
    write_report(report_path, source_url, "failure", result, notes, errors, snapshot_path, output_file)
    logging.error("Sync failed. Report: %s", report_path)
    return 1


if __name__ == "__main__":
    raise SystemExit(run())
