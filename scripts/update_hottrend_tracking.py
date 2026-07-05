from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
DEFAULT_SCORES = ROOT / "data" / "topic_scores.json"
DEFAULT_HISTORY = ROOT / "data" / "hottrend_topic_history.csv"
DEFAULT_EXCEL = ROOT / "data" / "hottrend_topic_history.xlsx"
DEFAULT_MASTER_EXCEL = ROOT / "data" / "master_dashboard.xlsx"
DEFAULT_LATEST = ROOT / "data" / "hottrend_latest_dashboard.csv"
DEFAULT_WEEKLY = ROOT / "data" / "hottrend_weekly_summary.csv"
DEFAULT_MONTHLY = ROOT / "data" / "hottrend_monthly_summary.csv"
DEFAULT_HTML = ROOT / "data" / "hottrend_dashboard.html"
BASE_URL = "https://smileaireviewhub.com"

from modules.performance_tracking import clean_excel_value, clean_sheet_name, print_workbook_report, safe_save_workbook


HISTORY_FIELDS = [
    "run_timestamp",
    "run_date",
    "run_week",
    "run_month",
    "topic",
    "slug",
    "normalized_key",
    "source",
    "trend_score",
    "seo_score",
    "traffic_score",
    "revenue_score",
    "buyer_intent_score",
    "video_priority",
    "social_score",
    "total_score",
    "score_grade",
    "recommendation",
    "priority",
    "content_decision",
    "status",
    "first_seen_date",
    "last_seen_date",
    "times_seen",
    "score_change_vs_previous",
    "rank_today",
    "rank_change_vs_previous",
    "trend_direction",
    "article_status",
    "video_status",
    "youtube_status",
    "index_status",
    "affiliate_status",
    "next_action",
    "article_url",
    "youtube_url",
    "notes",
]

SUMMARY_FIELDS = [
    "period",
    "topic",
    "slug",
    "source",
    "average_score",
    "best_score",
    "latest_score",
    "times_seen",
    "recommendation_trend",
    "video_priority_trend",
    "latest_status",
    "article_url",
    "youtube_url",
]

LATEST_FIELDS = [
    "rank_today",
    "topic",
    "slug",
    "source",
    "total_score",
    "score_grade",
    "recommendation",
    "priority",
    "trend_score",
    "seo_score",
    "traffic_score",
    "revenue_score",
    "buyer_intent_score",
    "video_priority",
    "social_priority",
    "content_decision",
    "first_seen_date",
    "last_seen_date",
    "times_seen",
    "score_change",
    "rank_change",
    "trend_direction",
    "article_status",
    "video_status",
    "youtube_status",
    "index_status",
    "affiliate_status",
    "next_action",
    "social_score",
    "recommended_action",
    "status",
    "article_url",
    "youtube_url",
    "notes",
]


def slugify(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower())
    return normalized.strip("-") or "untitled"


def today_parts(now: datetime | None = None) -> tuple[str, str, str, str]:
    current = now or datetime.now(timezone.utc)
    iso = current.astimezone(timezone.utc).isoformat(timespec="seconds")
    day = current.date().isoformat()
    iso_year, iso_week, _ = current.isocalendar()
    week = f"{iso_year}-W{iso_week:02d}"
    month = f"{current.year:04d}-{current.month:02d}"
    return iso, day, week, month


def read_json(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("scores", "topics", "selected_topics"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
    raise ValueError(f"Unsupported score JSON format: {path}")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def row_identity(row: dict[str, Any]) -> str:
    return str(row.get("normalized_key") or slugify(row.get("slug") or row.get("topic") or "untitled-topic"))


def write_csv_rows(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def numeric(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def social_score(row: dict[str, Any]) -> float:
    scores = row.get("social_scores")
    if not isinstance(scores, dict) or not scores:
        return numeric(row.get("social_score"))
    values = [numeric(value) for value in scores.values()]
    return round(sum(values) / max(len(values), 1), 1)


def score_grade(total_score: Any) -> str:
    score = numeric(total_score)
    if score >= 90:
        return "Excellent"
    if score >= 80:
        return "Strong"
    if score >= 70:
        return "Good"
    if score >= 60:
        return "Watch"
    return "Skip"


def priority_for(total_score: Any, recommendation: Any) -> str:
    text = str(recommendation or "").lower()
    score = numeric(total_score)
    if score >= 90 or "excellent" in text:
        return "P1"
    if score >= 80 or "strong" in text:
        return "P2"
    if score >= 70 or "good" in text:
        return "P3"
    if score >= 60 or "watch" in text:
        return "Watch"
    return "Skip"


def social_priority(row: dict[str, Any]) -> str:
    score = social_score(row)
    if score >= 75:
        return "High"
    if score >= 60:
        return "Medium"
    if score >= 45:
        return "Low"
    return "None"


def trend_direction(score_change: Any, rank_change: Any) -> str:
    score_delta = numeric(score_change)
    rank_delta = numeric(rank_change)
    if score_delta >= 5 or rank_delta >= 2:
        return "Rising"
    if score_delta <= -5 or rank_delta <= -2:
        return "Declining"
    if score_change == "" and rank_change == "":
        return "New"
    return "Stable"


def status_for(row: dict[str, Any]) -> str:
    recommendation = str(row.get("recommendation") or "").lower()
    decision = str(row.get("content_decision") or "").lower()
    if "skip" in recommendation or "skip" in decision:
        return "SKIP"
    if "excellent" in recommendation or "strong" in recommendation:
        return "PRIORITY"
    if "good" in recommendation or "article" in decision or "youtube" in decision or "website" in decision:
        return "CANDIDATE"
    return "MONITOR"


def recommended_action(row: dict[str, Any]) -> str:
    recommendation = str(row.get("recommendation") or "").lower()
    decision = str(row.get("content_decision") or "").lower()
    video = str(row.get("video_priority") or "").lower()
    social = social_score(row)
    if "skip" in recommendation or "skip" in decision:
        return "Skip"
    if "website" in decision or "article" in decision or numeric(row.get("seo_score")) >= 70:
        return "Article"
    if "youtube" in decision or video not in {"", "no video", "none"}:
        return "Video"
    if social >= 65:
        return "Social"
    if numeric(row.get("quora_potential")) >= 60:
        return "Quora"
    return "Monitor"


def article_status(row: dict[str, Any]) -> str:
    decision = str(row.get("content_decision") or "").lower()
    url = str(row.get("article_url") or "").strip()
    if url:
        return "Planned"
    if "article" in decision or "website" in decision:
        return "Candidate"
    return "Monitor"


def video_status(row: dict[str, Any]) -> str:
    video = str(row.get("video_priority") or "").strip()
    decision = str(row.get("content_decision") or "").lower()
    if video and video.lower() not in {"no video", "none"}:
        return "Candidate"
    if "youtube" in decision or "video" in decision:
        return "Candidate"
    return "Monitor"


def youtube_status(row: dict[str, Any]) -> str:
    return "Linked" if str(row.get("youtube_url") or "").strip() else "Not linked"


def index_status(row: dict[str, Any]) -> str:
    return "Needs article" if not str(row.get("article_url") or "").strip() else "Ready after publish"


def affiliate_status(row: dict[str, Any]) -> str:
    if numeric(row.get("revenue_score")) >= 65 or numeric(row.get("buyer_intent_score")) >= 65:
        return "High potential"
    if numeric(row.get("revenue_score")) >= 50:
        return "Medium potential"
    return "Low/unknown"


def trend_value(values: list[str]) -> str:
    cleaned = [value for value in values if value]
    if not cleaned:
        return ""
    if len(cleaned) == 1:
        return cleaned[0]
    return " -> ".join(cleaned[-3:])


def normalize_score_row(row: dict[str, Any], rank: int, run: tuple[str, str, str, str], previous_rows: list[dict[str, str]]) -> dict[str, Any]:
    run_timestamp, run_date, run_week, run_month = run
    topic = str(row.get("topic") or "").strip()
    slug = str(row.get("slug") or "").strip() or slugify(topic)
    key = slugify(slug or topic)
    previous_for_topic = [item for item in previous_rows if row_identity(item) == key]
    previous = previous_for_topic[-1] if previous_for_topic else {}
    previous_score = numeric(previous.get("total_score"), default=numeric(row.get("total_score")))
    previous_rank = numeric(previous.get("rank_today"), default=rank)
    total = numeric(row.get("total_score"))
    first_seen = previous_for_topic[0].get("first_seen_date") if previous_for_topic else run_date
    times_seen = len(previous_for_topic) + 1
    article_url = str(row.get("article_url") or "").strip() or f"{BASE_URL}/{slug}/"
    score_change = round(total - previous_score, 1) if previous else ""
    rank_change = int(previous_rank - rank) if previous else ""
    normalized = {
        "run_timestamp": run_timestamp,
        "run_date": run_date,
        "run_week": run_week,
        "run_month": run_month,
        "topic": topic,
        "slug": slug,
        "normalized_key": key,
        "source": row.get("source", ""),
        "trend_score": row.get("trend_score", ""),
        "seo_score": row.get("seo_score", ""),
        "traffic_score": row.get("traffic_score", ""),
        "revenue_score": row.get("revenue_score", ""),
        "buyer_intent_score": row.get("buyer_intent", row.get("buyer_intent_score", "")),
        "video_priority": row.get("video_priority", ""),
        "social_score": social_score(row),
        "total_score": total,
        "score_grade": row.get("score_grade") or score_grade(total),
        "recommendation": row.get("recommendation", ""),
        "priority": priority_for(total, row.get("recommendation", "")),
        "content_decision": row.get("content_decision", ""),
        "status": status_for(row),
        "first_seen_date": first_seen,
        "last_seen_date": run_date,
        "times_seen": times_seen,
        "score_change_vs_previous": score_change,
        "score_change": score_change,
        "rank_today": rank,
        "rank_change_vs_previous": rank_change,
        "rank_change": rank_change,
        "article_url": article_url,
        "youtube_url": row.get("youtube_url", ""),
        "notes": row.get("reason", ""),
    }
    normalized["recommended_action"] = recommended_action({**row, **normalized})
    normalized["next_action"] = normalized["recommended_action"]
    normalized["social_priority"] = social_priority({**row, **normalized})
    normalized["trend_direction"] = trend_direction(score_change, rank_change)
    normalized["article_status"] = article_status(normalized)
    normalized["video_status"] = video_status(normalized)
    normalized["youtube_status"] = youtube_status(normalized)
    normalized["index_status"] = index_status(normalized)
    normalized["affiliate_status"] = affiliate_status(normalized)
    return normalized


def sort_scores(scores: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(scores, key=lambda row: (-numeric(row.get("total_score")), str(row.get("topic", ""))))


def build_summary(history: list[dict[str, Any]], period_field: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in history:
        grouped[(str(row.get(period_field, "")), row_identity(row))].append(row)
    summary: list[dict[str, Any]] = []
    for (period, _), rows in grouped.items():
        rows = sorted(rows, key=lambda row: str(row.get("run_timestamp", "")))
        scores = [numeric(row.get("total_score")) for row in rows]
        latest = rows[-1]
        summary.append(
            {
                "period": period,
                "topic": latest.get("topic", ""),
                "slug": latest.get("slug", ""),
                "source": latest.get("source", ""),
                "average_score": round(sum(scores) / max(len(scores), 1), 1),
                "best_score": round(max(scores) if scores else 0, 1),
                "latest_score": round(scores[-1] if scores else 0, 1),
                "times_seen": len(rows),
                "recommendation_trend": trend_value([str(row.get("recommendation", "")) for row in rows]),
                "video_priority_trend": trend_value([str(row.get("video_priority", "")) for row in rows]),
                "latest_status": latest.get("status", ""),
                "article_url": latest.get("article_url", ""),
                "youtube_url": latest.get("youtube_url", ""),
            }
        )
    return sorted(summary, key=lambda row: (str(row.get("period", "")), -numeric(row.get("latest_score")), str(row.get("topic", ""))))


def latest_run_rows(history: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not history:
        return []
    latest_timestamp = max(str(row.get("run_timestamp", "")) for row in history)
    rows = [row for row in history if str(row.get("run_timestamp", "")) == latest_timestamp]
    return sorted(rows, key=lambda row: (numeric(row.get("rank_today")), -numeric(row.get("total_score"))))


def build_latest_dashboard(history: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in latest_run_rows(history):
        rows.append({**row, "recommended_action": recommended_action(row)})
    return rows


def write_excel_workbook(
    excel_path: Path,
    latest: list[dict[str, Any]],
    history: list[dict[str, Any]],
    weekly: list[dict[str, Any]],
    monthly: list[dict[str, Any]],
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
    except ImportError:
        return False

    workbook = Workbook()
    workbook.remove(workbook.active)
    top_revenue = sorted(latest, key=lambda row: numeric(row.get("revenue_score")), reverse=True)
    top_seo = sorted(latest, key=lambda row: numeric(row.get("seo_score")), reverse=True)
    top_video = sorted(latest, key=lambda row: (row.get("video_status") == "Candidate", numeric(row.get("total_score"))), reverse=True)
    top_social = sorted(latest, key=lambda row: numeric(row.get("social_score")), reverse=True)
    evergreen = sorted(history, key=lambda row: (numeric(row.get("total_score")), str(row.get("last_seen_date", ""))), reverse=True)
    trending_today = [row for row in latest if row.get("trend_direction") in {"New", "Rising", "Stable"}]
    latest_week = max((str(row.get("run_week", "")) for row in history), default="")
    latest_month = max((str(row.get("run_month", "")) for row in history), default="")
    sheets = [
        ("Dashboard", latest[:25], LATEST_FIELDS),
        ("Hottrend Latest", latest, LATEST_FIELDS),
        ("Today's Winners", [row for row in latest if numeric(row.get("total_score")) >= 70], LATEST_FIELDS),
        ("Top Revenue", top_revenue[:25], LATEST_FIELDS),
        ("Top SEO", top_seo[:25], LATEST_FIELDS),
        ("Top Video", top_video[:25], LATEST_FIELDS),
        ("Top Social", top_social[:25], LATEST_FIELDS),
        ("Evergreen", evergreen[:25], HISTORY_FIELDS),
        ("Trending Today", trending_today[:25], LATEST_FIELDS),
        ("Trending This Week", [row for row in weekly if row.get("period") == latest_week], SUMMARY_FIELDS),
        ("Trending This Month", [row for row in monthly if row.get("period") == latest_month], SUMMARY_FIELDS),
        ("Full History", history, HISTORY_FIELDS),
        ("Weekly Summary", weekly, SUMMARY_FIELDS),
        ("Monthly Summary", monthly, SUMMARY_FIELDS),
        ("Rising Topics", [row for row in latest if row.get("trend_direction") == "Rising"], LATEST_FIELDS),
        ("Declining Topics", [row for row in latest if row.get("trend_direction") == "Declining"], LATEST_FIELDS),
        ("Rejected Topics", [row for row in latest if row.get("recommended_action") == "Skip" or row.get("status") == "SKIP"], LATEST_FIELDS),
        ("Video Candidates", [row for row in latest if row.get("recommended_action") == "Video"], LATEST_FIELDS),
        ("Article Candidates", [row for row in latest if row.get("recommended_action") == "Article"], LATEST_FIELDS),
        ("Affiliate Candidates", [row for row in latest if str(row.get("affiliate_status", "")).startswith("High")], LATEST_FIELDS),
        ("Monitor Skip", [row for row in latest if row.get("recommended_action") in {"Monitor", "Skip"}], LATEST_FIELDS),
        ("Index Status", latest, LATEST_FIELDS),
        ("YouTube Status", latest, LATEST_FIELDS),
    ]
    used_sheet_names: set[str] = set()
    for title, rows, fields in sheets:
        ws = workbook.create_sheet(clean_sheet_name(title, used_sheet_names))
        ws.append([clean_excel_value(field) for field in fields])
        for row in rows:
            ws.append([clean_excel_value(row.get(field, "")) for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            max_width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_width + 2, 10), 60)
        if "score_change_vs_previous" in fields and ws.max_row > 1:
            col = fields.index("score_change_vs_previous") + 1
            letter = ws.cell(row=1, column=col).column_letter
            green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=green))
            ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=red))
        if "priority" in fields and ws.max_row > 1:
            col = fields.index("priority") + 1
            letter = ws.cell(row=1, column=col).column_letter
            yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="equal", formula=['"P1"'], fill=green))
            ws.conditional_formatting.add(f"{letter}2:{letter}{ws.max_row}", CellIsRule(operator="equal", formula=['"P2"'], fill=yellow))
    result = safe_save_workbook(workbook, excel_path)
    print_workbook_report(result)
    return bool(result.get("saved") or result.get("pending"))


def html_escape(value: Any) -> str:
    return (
        str(value or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def write_html_dashboard(path: Path, latest: list[dict[str, Any]]) -> None:
    top = latest[:10]
    video_candidates = [row for row in latest if row.get("recommended_action") == "Video"][:10]
    article_candidates = [row for row in latest if row.get("recommended_action") == "Article"][:10]
    affiliate_candidates = [row for row in latest if str(row.get("affiliate_status", "")).startswith("High")][:10]

    def table(rows: list[dict[str, Any]], fields: list[str]) -> str:
        header = "".join(f"<th>{html_escape(field.replace('_', ' ').title())}</th>" for field in fields)
        body = []
        for row in rows:
            cells = "".join(f"<td>{html_escape(row.get(field, ''))}</td>" for field in fields)
            body.append(f"<tr>{cells}</tr>")
        return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body)}</tbody></table>"

    fields = ["rank_today", "topic", "total_score", "score_grade", "recommendation", "priority", "trend_direction", "next_action"]
    html = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Hottrend Dashboard</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 32px; color: #172033; background: #f7f9fc; }}
    h1, h2 {{ color: #0f766e; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0 32px; background: #fff; }}
    th, td {{ border: 1px solid #d9e2ec; padding: 8px 10px; text-align: left; vertical-align: top; }}
    th {{ background: #e6fffb; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
  </style>
</head>
<body>
  <h1>Hottrend Dashboard</h1>
  <p>Local review dashboard. Do not publish unless explicitly requested.</p>
  <h2>Top 10 Topics</h2>
  {table(top, fields)}
  <h2>Video Candidates</h2>
  {table(video_candidates, fields)}
  <h2>Article Candidates</h2>
  {table(article_candidates, fields)}
  <h2>Affiliate Candidates</h2>
  {table(affiliate_candidates, fields + ["affiliate_status"])}
</body>
</html>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")


def update_hottrend_tracking(
    scores_path: Path = DEFAULT_SCORES,
    history_path: Path = DEFAULT_HISTORY,
    excel_path: Path = DEFAULT_EXCEL,
    master_excel_path: Path = DEFAULT_MASTER_EXCEL,
    latest_dashboard_path: Path = DEFAULT_LATEST,
    weekly_path: Path = DEFAULT_WEEKLY,
    monthly_path: Path = DEFAULT_MONTHLY,
    html_dashboard_path: Path = DEFAULT_HTML,
    now: datetime | None = None,
) -> dict[str, Any]:
    scores = sort_scores(read_json(scores_path))
    existing = read_csv_rows(history_path)
    run = today_parts(now)
    new_rows = [normalize_score_row(row, index, run, existing) for index, row in enumerate(scores, 1)]
    history = [*existing, *new_rows]
    weekly = build_summary(history, "run_week")
    monthly = build_summary(history, "run_month")
    latest = build_latest_dashboard(history)

    write_csv_rows(history_path, history, HISTORY_FIELDS)
    write_csv_rows(weekly_path, weekly, SUMMARY_FIELDS)
    write_csv_rows(monthly_path, monthly, SUMMARY_FIELDS)
    write_csv_rows(latest_dashboard_path, latest, LATEST_FIELDS)
    excel_written = write_excel_workbook(excel_path, latest, history, weekly, monthly)
    master_excel_written = write_excel_workbook(master_excel_path, latest, history, weekly, monthly)
    write_html_dashboard(html_dashboard_path, latest)
    business_outputs = {}
    try:
        from modules.performance_tracking import write_business_outputs

        business_outputs = write_business_outputs()
    except Exception as exc:  # Reporting should not break topic scoring.
        business_outputs = {"error": str(exc)}
    return {
        "topics": len(new_rows),
        "history_rows": len(history),
        "history": str(history_path),
        "latest_dashboard": str(latest_dashboard_path),
        "weekly": str(weekly_path),
        "monthly": str(monthly_path),
        "excel": str(excel_path),
        "excel_written": excel_written,
        "master_excel": str(master_excel_path),
        "master_excel_written": master_excel_written,
        "html_dashboard": str(html_dashboard_path),
        "business_outputs": business_outputs,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Append scored hottrend topics to persistent history CSV/Excel dashboards.")
    parser.add_argument("--scores", default=str(DEFAULT_SCORES), help="Input topic score JSON path.")
    parser.add_argument("--history", default=str(DEFAULT_HISTORY), help="Output history CSV path.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Output Excel workbook path.")
    parser.add_argument("--master-excel", default=str(DEFAULT_MASTER_EXCEL), help="Output master Excel workbook path.")
    parser.add_argument("--latest-dashboard", default=str(DEFAULT_LATEST), help="Output latest dashboard CSV path.")
    parser.add_argument("--weekly", default=str(DEFAULT_WEEKLY), help="Output weekly summary CSV path.")
    parser.add_argument("--monthly", default=str(DEFAULT_MONTHLY), help="Output monthly summary CSV path.")
    parser.add_argument("--html-dashboard", default=str(DEFAULT_HTML), help="Output local HTML dashboard path.")
    args = parser.parse_args()

    result = update_hottrend_tracking(
        scores_path=Path(args.scores),
        history_path=Path(args.history),
        excel_path=Path(args.excel),
        master_excel_path=Path(args.master_excel),
        latest_dashboard_path=Path(args.latest_dashboard),
        weekly_path=Path(args.weekly),
        monthly_path=Path(args.monthly),
        html_dashboard_path=Path(args.html_dashboard),
    )
    print(f"Hottrend tracking updated: topics={result['topics']} history_rows={result['history_rows']}")
    print(f"History CSV: {result['history']}")
    print(f"Latest dashboard CSV: {result['latest_dashboard']}")
    print(f"Weekly summary CSV: {result['weekly']}")
    print(f"Monthly summary CSV: {result['monthly']}")
    if result["excel_written"]:
        print(f"Excel workbook: {result['excel']}")
    else:
        print("Excel workbook skipped: openpyxl is not available.")
    if result["master_excel_written"]:
        print(f"Master Excel workbook: {result['master_excel']}")
    else:
        print("Master Excel workbook skipped: openpyxl is not available.")
    print(f"HTML dashboard: {result['html_dashboard']}")
    if result.get("business_outputs"):
        print(f"Business dashboard outputs: {result['business_outputs']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
