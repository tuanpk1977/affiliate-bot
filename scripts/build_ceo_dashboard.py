from __future__ import annotations

import html
import json
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import settings  # noqa: E402


DATA_DIR = settings.data_dir
HTML_PATH = DATA_DIR / "daily_ceo_dashboard.html"
JSON_PATH = DATA_DIR / "daily_ceo_dashboard.json"
XLSX_PATH = DATA_DIR / "master_dashboard.xlsx"


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def build_dashboard_payload() -> dict[str, Any]:
    weekly_topics = _read_json(DATA_DIR / "weekly_topics.json", [])
    editorial_calendar = _read_json(DATA_DIR / "editorial_calendar.json", [])
    content_review = _read_json(DATA_DIR / "content_review_report.json", {"summary": {}, "items": []})
    publish_gate = _read_json(DATA_DIR / "publish_gate_report.json", {"summary": {}, "items": []})
    optimization = _read_json(DATA_DIR / "optimization_report.json", {"actions": []})
    research_quality = _read_json(DATA_DIR / "research_quality_report.json", [])
    enrichment_queue = _read_json(DATA_DIR / "research_enrichment_queue.json", [])
    human_approval = _read_json(DATA_DIR / "human_approval_queue.json", [])
    publish_queue = _read_json(DATA_DIR / "publish_queue.json", [])

    monday_rows = [row for row in editorial_calendar if str(row.get("day_of_week", "")) == "Monday"]
    approved_monday = [row for row in publish_queue if str(row.get("status", "")) == "published_local"]
    blocked_publish = [row for row in publish_queue if str(row.get("status", "")) == "blocked"]
    needs_human = [row for row in human_approval if str(row.get("status", "")) == "needs_human_review"]
    needs_enrichment = [row for row in enrichment_queue if str(row.get("status", "")) in {"pending", "needs_enrichment", "enriching"}]

    payload = {
        "generated_at": datetime.now(UTC).isoformat(),
        "summary": {
            "weekly_topics": len(weekly_topics),
            "monday_calendar_articles": len(monday_rows),
            "published_local": len(approved_monday),
            "publish_blocked": len(blocked_publish),
            "needs_human_review": len(needs_human),
            "needs_enrichment": len(needs_enrichment),
            "content_review_items": len(content_review.get("items", [])),
            "publish_queue_items": len(publish_queue),
            "optimization_actions": len(optimization.get("actions", [])),
        },
        "weekly_topics": weekly_topics[:10],
        "monday_calendar": monday_rows[:10],
        "blocked_publish": blocked_publish[:10],
        "needs_human_review": needs_human[:10],
        "needs_enrichment": needs_enrichment[:10],
        "optimization_actions": optimization.get("actions", [])[:10],
        "research_quality": research_quality[:10],
    }
    return payload


def write_html(payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Daily CEO Dashboard</title>
  <style>
    body {{ font-family: Georgia, 'Segoe UI', serif; margin: 0; background: linear-gradient(180deg, #f5f0e8, #fffdf8); color: #1f1b16; }}
    .wrap {{ max-width: 1100px; margin: 0 auto; padding: 32px 20px 56px; }}
    .hero {{ background: #1f3a5f; color: #fff8ef; padding: 24px; border-radius: 18px; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 14px; margin: 20px 0 28px; }}
    .card {{ background: white; border: 1px solid #eadfce; border-radius: 16px; padding: 18px; box-shadow: 0 8px 30px rgba(31,27,22,0.06); }}
    h1,h2 {{ margin-top: 0; }}
    ul {{ padding-left: 20px; }}
    code {{ background: #f3ede3; padding: 1px 5px; border-radius: 6px; }}
  </style>
</head>
<body>
  <main class="wrap">
    <section class="hero">
      <p>Local-only operational dashboard</p>
      <h1>Weekly Content Production Dashboard</h1>
      <p>Generated at {html.escape(payload['generated_at'])}. No deploy, no push, no IndexNow.</p>
    </section>
    <section class="grid">
      <div class="card"><h2>{summary['weekly_topics']}</h2><p>Weekly topics</p></div>
      <div class="card"><h2>{summary['monday_calendar_articles']}</h2><p>Monday calendar rows</p></div>
      <div class="card"><h2>{summary['published_local']}</h2><p>Published locally</p></div>
      <div class="card"><h2>{summary['publish_blocked']}</h2><p>Publish blocked</p></div>
      <div class="card"><h2>{summary['needs_human_review']}</h2><p>Needs human review</p></div>
      <div class="card"><h2>{summary['needs_enrichment']}</h2><p>Needs enrichment</p></div>
    </section>
    <section class="card"><h2>Monday topics</h2><ul>{''.join(f"<li><strong>{html.escape(str(row.get('keyword','')))}</strong> - {html.escape(str(row.get('article_type','')))}</li>" for row in payload['monday_calendar']) or '<li>None</li>'}</ul></section>
    <section class="card"><h2>Blocked publish items</h2><ul>{''.join(f"<li><code>{html.escape(str(row.get('slug','')))}</code>: {html.escape('; '.join(row.get('failures', [])) if isinstance(row.get('failures'), list) else str(row.get('failures','')))}</li>" for row in payload['blocked_publish']) or '<li>None</li>'}</ul></section>
    <section class="card"><h2>Needs human review</h2><ul>{''.join(f"<li><code>{html.escape(str(row.get('slug','')))}</code>: {html.escape(str(row.get('topic','')))}</li>" for row in payload['needs_human_review']) or '<li>None</li>'}</ul></section>
    <section class="card"><h2>Optimization actions</h2><ul>{''.join(f"<li><code>{html.escape(str(row.get('slug','')))}</code>: {html.escape(str(row.get('next_recommended_action','')))}</li>" for row in payload['optimization_actions']) or '<li>None</li>'}</ul></section>
  </main>
</body>
</html>
"""
    HTML_PATH.write_text(html_text, encoding="utf-8")


def write_workbook(payload: dict[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in payload["summary"].items():
        summary_sheet.append([key, value])

    def add_sheet(name: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(title=name[:31])
        if not rows:
            sheet.append(["empty"])
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([json.dumps(row.get(header), ensure_ascii=False) if isinstance(row.get(header), (list, dict)) else row.get(header) for header in headers])

    add_sheet("WeeklyTopics", payload["weekly_topics"])
    add_sheet("MondayCalendar", payload["monday_calendar"])
    add_sheet("BlockedPublish", payload["blocked_publish"])
    add_sheet("NeedsHuman", payload["needs_human_review"])
    add_sheet("NeedsEnrichment", payload["needs_enrichment"])
    add_sheet("Optimization", payload["optimization_actions"])
    add_sheet("ResearchQuality", payload["research_quality"])
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(XLSX_PATH)


def main(*, verbose: bool = True) -> int:
    payload = build_dashboard_payload()
    _write_json(JSON_PATH, payload)
    write_html(payload)
    write_workbook(payload)
    if verbose:
        print(json.dumps({"daily_ceo_dashboard_html": str(HTML_PATH), "master_dashboard_xlsx": str(XLSX_PATH), "summary": payload["summary"]}, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
